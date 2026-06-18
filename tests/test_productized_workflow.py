from datetime import date
from pathlib import Path
from contextlib import closing
import sqlite3
import subprocess
from tempfile import TemporaryDirectory
import unittest
from io import BytesIO
from zipfile import ZipFile
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from ops_data_workflow.ai import _build_payload, group_topic_labels, resolve_deepseek_settings
from ops_data_workflow.content_ledger import LEDGER_COLUMNS
from ops_data_workflow.categories import CATEGORY_TAG_MAP, category_from_tags
from ops_data_workflow.source_channels import infer_channel_from_path
from ops_data_workflow.analysis_jobs import list_analysis_jobs
from ops_data_workflow.storage import init_db
from ops_data_workflow.storage import (
    purge_history_state,
    read_batch_record,
)
from ops_data_workflow.upload_input import (
    detect_upload_channel_conflicts,
    infer_period_from_upload_names,
    materialize_uploaded_files,
)
from ops_data_workflow.workflow import run_archived_workflow
from tests.test_workflow import _write_raw_fixture


class FakeUpload:
    def __init__(self, name: str, data: bytes):
        self.name = name
        self._data = data

    def getvalue(self) -> bytes:
        return self._data


def _workbook_bytes(frame: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="sheet1", index=False)
    return buffer.getvalue()


def _write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


class ProductizedWorkflowTests(unittest.TestCase):
    def test_materialize_uploaded_files_accepts_direct_excel_csv_and_zip(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            zip_buffer = BytesIO()
            with ZipFile(zip_buffer, "w") as archive:
                archive.writestr("抖音商业化.csv", "视频标题,视频id,素材ID,消耗,展示数,点击数,激活数,付费次数,内容类型\n标题,dy,mat,1,2,1,1,1,热点行情\n")

            uploads = [
                FakeUpload(
                    "B站.xlsx",
                    _workbook_bytes(
                        pd.DataFrame(
                            [
                                {
                                    "视频AVID": "av",
                                    "视频BVID": "bv",
                                    "视频标题": "标题",
                                    "花费": 1,
                                    "展示量": 2,
                                    "点击量": 1,
                                    "应用激活数": 1,
                                    "应用内付费": 1,
                                }
                            ]
                        )
                    ),
                ),
                FakeUpload(
                    "小红书商业化.csv",
                    "标题,笔记ID,发布作者,类型,内容分类,消费,展现量,点击量,激活数,首次付费次数\n标题,note,作者,图文,热点行情,1,2,1,1,1\n".encode(
                        "utf-8-sig"
                    ),
                ),
                FakeUpload("raw.zip", zip_buffer.getvalue()),
            ]

            result = materialize_uploaded_files(uploads, tmp_path / "raw")

            self.assertTrue((result.raw_dir / "B站.xlsx").exists())
            self.assertTrue((result.raw_dir / "小红书商业化.csv").exists())
            self.assertTrue((result.raw_dir / "抖音商业化.csv").exists())
            self.assertFalse((tmp_path / "uploaded_originals").exists())
            self.assertEqual(len(result.original_files), 2)

    def test_materialize_uploaded_files_keeps_clean_named_uploads_as_sources(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            result = materialize_uploaded_files(
                [
                    FakeUpload(
                        "小红书市场部_clean.xlsx",
                        _workbook_bytes(pd.DataFrame([{"周期": "2026-05-19 至 2026-05-25"}])),
                    ),
                    FakeUpload(
                        "小红书市场部.xlsx",
                        _workbook_bytes(pd.DataFrame([{"笔记/素材ID": "note-1", "消费": 10, "激活数": 1}])),
                    ),
                ],
                tmp_path / "raw",
            )

            self.assertTrue((result.raw_dir / "小红书市场部_clean.xlsx").exists())
            self.assertTrue((result.raw_dir / "小红书市场部.xlsx").exists())
            self.assertEqual(
                [path.name for path in result.original_files],
                ["小红书市场部_clean.xlsx", "小红书市场部.xlsx"],
            )

    def test_materialize_uploaded_files_replaces_same_channel_and_keeps_other_channels(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw" / "20260401-20260407"
            raw_dir.mkdir(parents=True)
            (raw_dir / "抖音商业化旧文件.csv").write_text(
                "视频标题,视频id,素材ID,消耗,展示数,点击数,激活数,付费次数,内容类型\n旧标题,old,old-mat,1,2,1,1,0,资讯\n",
                encoding="utf-8-sig",
            )
            (raw_dir / "B站.csv").write_text(
                "视频BVID,视频标题,花费,展示量,点击量,应用激活数,应用内付费\nBV1,B站标题,10,100,10,1,0\n",
                encoding="utf-8-sig",
            )

            materialize_uploaded_files(
                [
                    FakeUpload(
                        "抖音商业化新文件.csv",
                        "视频标题,视频id,素材ID,消耗,展示数,点击数,激活数,付费次数,内容类型\n新标题,new,new-mat,9,20,10,3,1,股友说\n".encode(
                            "utf-8-sig"
                        ),
                    )
                ],
                raw_dir,
                replace_same_channel=True,
            )

            self.assertFalse((raw_dir / "抖音商业化旧文件.csv").exists())
            self.assertTrue((raw_dir / "抖音商业化新文件.csv").exists())
            self.assertTrue((raw_dir / "B站.csv").exists())

    def test_materialize_uploaded_files_rejects_existing_channel_by_default(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw" / "20260401-20260407"
            raw_dir.mkdir(parents=True)
            existing = raw_dir / "抖音商业化旧文件.csv"
            existing.write_text(
                "视频标题,视频id,素材ID,消耗,展示数,点击数,激活数,付费次数,内容类型\n旧标题,old,old-mat,1,2,1,1,0,资讯\n",
                encoding="utf-8-sig",
            )
            other_channel = raw_dir / "B站.csv"
            other_channel.write_text(
                "视频BVID,视频标题,花费,展示量,点击量,应用激活数,应用内付费\nBV1,B站标题,10,100,10,1,0\n",
                encoding="utf-8-sig",
            )
            uploads = [
                FakeUpload(
                    "抖音商业化新文件.csv",
                    "视频标题,视频id,素材ID,消耗,展示数,点击数,激活数,付费次数,内容类型\n新标题,new,new-mat,9,20,10,3,1,股友说\n".encode(
                        "utf-8-sig"
                    ),
                )
            ]

            conflicts = detect_upload_channel_conflicts(uploads, raw_dir)
            with self.assertRaisesRegex(FileExistsError, "本地已存在渠道"):
                materialize_uploaded_files(uploads, raw_dir)

            self.assertEqual([conflict.channel for conflict in conflicts], ["抖音商业化"])
            self.assertTrue(existing.exists())
            self.assertTrue(other_channel.exists())
            self.assertFalse((raw_dir / "抖音商业化新文件.csv").exists())

    def test_social_uploads_replace_same_business_channel_across_wechat_tencent_video_account(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw" / "20260401-20260407"
            raw_dir.mkdir(parents=True)
            (raw_dir / "微信市场部旧文件.csv").write_text(
                "创意名称,花费,曝光次数,点击次数,APP激活次数,注册次数\n旧微信创意,1,2,1,1,0\n",
                encoding="utf-8-sig",
            )
            (raw_dir / "B站.csv").write_text(
                "视频BVID,视频标题,花费,展示量,点击量,应用激活数,应用内付费\nBV1,B站标题,10,100,10,1,0\n",
                encoding="utf-8-sig",
            )

            self.assertEqual(infer_channel_from_path("微信市场部.xlsx"), "微信市场部")
            self.assertEqual(infer_channel_from_path("腾讯（市场部）.xlsx"), "微信市场部")
            self.assertEqual(infer_channel_from_path("视频号投放.xlsx"), "微信市场部")
            self.assertEqual(infer_channel_from_path("微信投放.xlsx"), "微信市场部")
            self.assertEqual(infer_channel_from_path("腾讯商业化.xlsx"), "微信商业化")
            self.assertEqual(infer_channel_from_path("视频号商业化.xlsx"), "微信商业化")
            self.assertEqual(infer_channel_from_path("快手投放.xlsx"), "快手投放")

            materialize_uploaded_files(
                [
                    FakeUpload(
                        "视频号投放.csv",
                        "创意名称,花费,曝光次数,点击次数,APP激活次数,注册次数\n新视频号创意,9,20,10,3,1\n".encode(
                            "utf-8-sig"
                        ),
                    )
                ],
                raw_dir,
                replace_same_channel=True,
            )

            self.assertFalse((raw_dir / "微信市场部旧文件.csv").exists())
            self.assertTrue((raw_dir / "视频号投放.csv").exists())
            self.assertTrue((raw_dir / "B站.csv").exists())

    def test_materialize_uploaded_files_writes_to_period_raw_directory(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uploads = [
                FakeUpload(
                    "B站.xlsx",
                    _workbook_bytes(
                        pd.DataFrame(
                            [
                                {
                                    "视频AVID": "av",
                                    "视频BVID": "bv",
                                    "视频标题": "标题",
                                    "花费": 1,
                                    "展示量": 2,
                                    "点击量": 1,
                                    "应用激活数": 1,
                                    "应用内付费": 1,
                                }
                            ]
                        )
                    ),
                )
            ]

            result = materialize_uploaded_files(
                uploads,
                tmp_path / "data" / "raw" / "20260401-20260427",
            )

            self.assertEqual(result.raw_dir.name, "20260401-20260427")
            self.assertTrue((tmp_path / "data" / "raw" / "20260401-20260427" / "B站.xlsx").exists())

    def test_materialize_uploaded_files_preserves_folder_relative_paths(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uploads = [
                FakeUpload(
                    "20260401-20260427/B站.xlsx",
                    _workbook_bytes(
                        pd.DataFrame(
                            [
                                {
                                    "视频AVID": "av",
                                    "视频BVID": "bv",
                                    "视频标题": "标题",
                                    "花费": 1,
                                    "展示量": 2,
                                    "点击量": 1,
                                    "应用激活数": 1,
                                    "应用内付费": 1,
                                }
                            ]
                        )
                    ),
                )
            ]

            result = materialize_uploaded_files(uploads, tmp_path / "raw")

            self.assertTrue((result.raw_dir / "20260401-20260427" / "B站.xlsx").exists())
            self.assertTrue((result.original_files[0]).exists())

    def test_materialize_uploaded_files_can_strip_common_period_folder(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uploads = [
                FakeUpload(
                    "20260401-20260427/B站.xlsx",
                    _workbook_bytes(
                        pd.DataFrame(
                            [
                                {
                                    "视频AVID": "av",
                                    "视频BVID": "bv",
                                    "视频标题": "标题",
                                    "花费": 1,
                                    "展示量": 2,
                                    "点击量": 1,
                                    "应用激活数": 1,
                                    "应用内付费": 1,
                                }
                            ]
                        )
                    ),
                )
            ]

            result = materialize_uploaded_files(
                uploads,
                tmp_path / "data" / "raw" / "20260401-20260427",
                strip_common_period_root=True,
            )

            self.assertTrue((result.raw_dir / "B站.xlsx").exists())
            self.assertFalse((result.raw_dir / "20260401-20260427" / "B站.xlsx").exists())
            self.assertTrue((result.original_files[0]).exists())

    def test_materialize_uploaded_files_rejects_parent_escape_paths(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)

            with self.assertRaisesRegex(ValueError, "非法路径"):
                materialize_uploaded_files(
                    [FakeUpload("../evil.xlsx", b"bad")],
                    tmp_path / "raw",
                )

    def test_materialize_uploaded_files_rejects_absolute_paths(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)

            with self.assertRaisesRegex(ValueError, "非法路径"):
                materialize_uploaded_files(
                    [FakeUpload("/tmp/evil.xlsx", b"bad")],
                    tmp_path / "raw",
                )

    def test_infer_period_from_upload_names_reads_top_level_folder(self):
        uploads = [FakeUpload("20260401-20260427/B站.xlsx", b"")]

        inferred = infer_period_from_upload_names(uploads)

        self.assertEqual(inferred, (date(2026, 4, 1), date(2026, 4, 27)))

    def test_infer_period_from_upload_names_supports_dashed_folder_pattern(self):
        uploads = [FakeUpload("2026-04-01_2026-04-27/抖音商业化.csv", b"")]

        inferred = infer_period_from_upload_names(uploads)

        self.assertEqual(inferred, (date(2026, 4, 1), date(2026, 4, 27)))

    def test_infer_period_from_upload_names_ignores_non_folder_uploads(self):
        uploads = [FakeUpload("B站.xlsx", b"")]

        self.assertIsNone(infer_period_from_upload_names(uploads))

    def test_cli_infers_period_from_source_directory_name(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "data" / "months" / "202604"
            raw_dir.mkdir(parents=True)
            _write_raw_fixture(raw_dir)

            completed = subprocess.run(
                [
                    ".venv/bin/python",
                    "main.py",
                    "--input",
                    str(raw_dir),
                    "--output",
                    str(tmp_path / "outputs"),
                    "--processed-root",
                    str(tmp_path / "processed"),
                    "--db",
                    str(tmp_path / "workflow.sqlite3"),
                    "--env",
                    str(tmp_path / "missing.env"),
                ],
                cwd=Path(__file__).resolve().parents[1],
                text=True,
                capture_output=True,
                check=False,
            )

            self.assertEqual(completed.returncode, 0, completed.stderr)
            with closing(sqlite3.connect(tmp_path / "workflow.sqlite3")) as conn:
                period = conn.execute(
                    """
                    select period_start, period_end, period_level, period_key, data_start, data_end, source_type
                    from upload_batches
                    """
                ).fetchone()
            self.assertEqual(
                period,
                ("2026-04-01", "2026-04-30", "month", "2026-04", "2026-04-01", "2026-04-30", "upload"),
            )

    def test_resolve_deepseek_settings_reads_explicit_env_without_leaking_secret(self):
        with TemporaryDirectory() as tmp:
            env_path = Path(tmp) / "deepseek.env"
            env_path.write_text(
                "DEEPSEEK_API_KEY=sk-test-secret\nDEEPSEEK_BASE_URL=https://example.test\nDEEPSEEK_MODEL=deepseek-test\n",
                encoding="utf-8",
            )

            settings = resolve_deepseek_settings(env_path)

            self.assertTrue(settings.configured)
            self.assertEqual(settings.base_url, "https://example.test")
            self.assertEqual(settings.model, "deepseek-test")
            self.assertIn(str(env_path), settings.checked_paths)
            self.assertNotIn("sk-test-secret", settings.public_status)

    def test_ai_payload_serializes_nullable_float_frames(self):
        payload = _build_payload(
            pd.DataFrame([{"channel": "总计", "spend": 1.0}]),
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(),
            "",
            platform_summary=pd.DataFrame(),
            platform_category_summary=pd.DataFrame(
                {"platform": ["抖音"], "first_pay_rate": pd.Series([pd.NA], dtype="Float64")}
            ),
        )

        self.assertEqual(payload["channel_category_topic_summary"][0]["first_pay_rate"], "")

    def test_group_topic_labels_returns_empty_without_deepseek_key(self):
        with TemporaryDirectory() as tmp:
            frame = pd.DataFrame(
                [
                    {
                        "channel": "抖音商业化",
                        "title": "短线交易高手",
                        "content_id": "dy-1",
                        "material_id": "mat-1",
                        "category_l2": "股友说",
                        "category_l3": "",
                        "spend": 100.0,
                        "activations": 10.0,
                        "first_pay_count": 2.0,
                    }
                ]
            )

            labels = group_topic_labels(frame, env_path=Path(tmp) / "missing.env")

            self.assertEqual(labels, {})

    def test_tag_mapping_uses_exact_user_defined_hashtags(self):
        self.assertEqual(CATEGORY_TAG_MAP["#同花顺资讯"], "资讯")
        self.assertEqual(CATEGORY_TAG_MAP["#同顺图解"], "图文")
        self.assertEqual(CATEGORY_TAG_MAP["#问财问句"], "问财问句")
        self.assertEqual(CATEGORY_TAG_MAP["#同花顺股民话题"], "社区话题")

        self.assertEqual(category_from_tags("今天的内容 #同顺图解 #财经"), "图文")
        self.assertEqual(category_from_tags("给短线交易者的完美范例 #股友说"), "")
        self.assertEqual(category_from_tags("给短线交易者的完美范例 #同花顺股友说"), "股友说")

    def test_archived_workflow_persists_batch_files_and_core_recap(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=tmp_path / "workflow.sqlite3",
                env_path=tmp_path / "missing.env",
            )

            self.assertTrue(result.batch_id)
            self.assertTrue(result.archive_dir.exists())
            self.assertEqual(result.ai_summary, "")
            self.assertIn("无历史对比数据", result.comparison_note)
            self.assertTrue((result.archive_dir / "cleaned.xlsx").exists())
            self.assertFalse((raw_dir / "cleaned.xlsx").exists())
            output_dir = result.core_recap_xlsx.parent
            self.assertTrue((output_dir / "content_recap_core.xlsx").exists())
            workbook = load_workbook(result.core_recap_xlsx, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["清洗后素材表", "内容复盘表", "不可分析汇总", "匹配覆盖率", "已匹配账号类型分析", "未匹配归因"],
            )
            workbook.close()

            record = read_batch_record(tmp_path / "workflow.sqlite3", result.batch_id)
            self.assertEqual(record["batch_id"], result.batch_id)
            self.assertEqual(record["archive_dir"], str(result.archive_dir))
            self.assertEqual(record["output_dir"], str(output_dir))

            with closing(sqlite3.connect(tmp_path / "workflow.sqlite3")) as conn:
                batch_count = conn.execute("select count(*) from upload_batches").fetchone()[0]
                file_count = conn.execute("select count(*) from uploaded_files").fetchone()[0]
                item_count = conn.execute("select count(*) from canonical_items").fetchone()[0]
                ai_count = conn.execute("select count(*) from ai_reports").fetchone()[0]
                coverage_count = conn.execute("select count(*) from attribution_coverage_items").fetchone()[0]
                matched_count = conn.execute("select count(*) from matched_attribution_items").fetchone()[0]
                unmatched_count = conn.execute("select count(*) from unmatched_attribution_items").fetchone()[0]

            self.assertEqual(batch_count, 1)
            self.assertGreaterEqual(file_count, 4)
            self.assertEqual(item_count, len(result.canonical))
            self.assertEqual(ai_count, 0)
            self.assertEqual(coverage_count, 4)
            self.assertEqual(matched_count, len(result.matched_attribution))
            self.assertEqual(unmatched_count, len(result.unmatched_attribution))
            self.assertGreaterEqual(unmatched_count, 0)
            self.assertIn("spend_share_of_total", result.attribution_coverage.columns)

    def test_archived_workflow_materializes_cleaned_workbook_before_analysis(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_xlsx(
                raw_dir / "B站.xlsx",
                {
                    "Sheet2": pd.DataFrame(
                        [
                            {
                                "单元名称": "BV001",
                                "Up主mid": "1622777305",
                                "求和项:总花费": 10.0,
                                "求和项:应用激活数": 2,
                                "求和项:应用内首次付费次数": 1,
                            }
                        ]
                    ),
                    "Sheet1": pd.DataFrame(
                        [
                            {
                                "单元名称": "BV001",
                                "Up主mid": "1622777305",
                                "总花费": 10.0,
                                "应用激活数": 2,
                                "应用内首次付费次数": 1,
                            }
                        ]
                    ),
                },
            )

            result = run_archived_workflow(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=tmp_path / "workflow.sqlite3",
                env_path=tmp_path / "missing.env",
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertFalse((raw_dir / "cleaned.xlsx").exists())
            self.assertFalse((raw_dir / "period_manifest.json").exists())
            self.assertTrue((result.archive_dir / "cleaned.xlsx").exists())
            self.assertTrue((result.archive_dir / "period_manifest.json").exists())
            import_log = pd.read_excel(result.archive_dir / "cleaned.xlsx", sheet_name="导入日志")
            self.assertEqual(set(import_log["sheet_name"]), {"Sheet1", "Sheet2"})
            self.assertEqual(len(result.canonical), 1)
            self.assertAlmostEqual(result.canonical["spend"].sum(), 10.0)
            self.assertTrue(result.canonical["source_sheet"].replace("", pd.NA).notna().all())

    def test_archived_workflow_replaces_same_period_rows_in_place(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            first = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-30",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )
            _write_xlsx(
                raw_dir / "小红书商业化.xlsx",
                {
                    "sheet1": pd.DataFrame(
                        [
                            {
                                "标题": "第二次导入内容",
                                "笔记ID": "note-second",
                                "发布作者": "同花顺理财",
                                "类型": "图文",
                                "内容分类": "热点行情",
                                "消费": 9,
                                "展现量": 90,
                                "点击量": 9,
                                "激活数": 3,
                                "首次付费次数": 1,
                            }
                        ]
                    )
                },
            )

            second = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-30",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            self.assertEqual(second.batch_id, first.batch_id)
            with closing(sqlite3.connect(db_path)) as conn:
                batch_count = conn.execute("select count(*) from upload_batches").fetchone()[0]
                canonical_count = conn.execute("select count(*) from canonical_items").fetchone()[0]
                ai_count = conn.execute("select count(*) from ai_reports").fetchone()[0]
            self.assertEqual(batch_count, 1)
            self.assertEqual(canonical_count, len(second.canonical))
            self.assertEqual(ai_count, 0)

    def test_archived_workflow_persists_match_performance_and_feishu_snapshot_tables(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)
            feishu_ledger = pd.DataFrame(
                [
                    {
                        "platform": "小红书",
                        "content_id": "note-1",
                        "content_url": "https://www.xiaohongshu.com/explore/note-1",
                        "title": "存储芯片板块再度爆发",
                        "account": "同花顺投资",
                        "tags": "财经",
                        "raw_content_type": "图文",
                        "category_l1": "投教",
                        "category_l2": "方法论",
                        "bilibili_content_type": "",
                        "content_type": "方法论",
                        "published_date": "2026-04-01",
                        "source_file": "harvester_feishu",
                        "source_sheet": "小红书",
                        "source_row": 2,
                    }
                ],
                columns=LEDGER_COLUMNS,
            )
            feishu_ledger.attrs["feishu_snapshot"] = {
                "enabled": True,
                "fetched_at": "2026-06-15T00:00:00+00:00",
                "total_rows": 1,
                "platform_counts": {"小红书": 1},
                "sheet_row_counts": {"xhsSheet": 1},
                "field_completeness": {"content_id": 1.0},
                "warnings": [],
            }
            feishu_ledger.attrs["source_files"] = set()

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=feishu_ledger):
                result = run_archived_workflow(
                    raw_dir,
                    "2026-04-01",
                    "2026-04-30",
                    output_root=tmp_path / "outputs",
                    archive_root=tmp_path / "archive",
                    db_path=db_path,
                    env_path=tmp_path / "missing.env",
                )

            with closing(sqlite3.connect(db_path)) as conn:
                canonical_count = conn.execute(
                    "select count(*) from canonical_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                performance_count = conn.execute(
                    "select count(*) from content_performance_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                match_count = conn.execute(
                    "select count(*) from asset_match_results where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                snapshot_count = conn.execute(
                    "select count(*) from feishu_ledger_snapshots where batch_id = ? and total_rows = 1",
                    (result.batch_id,),
                ).fetchone()[0]

            self.assertEqual(canonical_count, len(result.canonical))
            self.assertGreater(performance_count, 0)
            self.assertGreater(match_count, 0)
            self.assertEqual(snapshot_count, 1)

    def test_archived_workflow_enqueues_top_multimodal_jobs_only_for_upload_triggers(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            upload_result = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-30",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=db_path,
                env_path=tmp_path / "missing.env",
                enqueue_background_analysis=True,
                background_trigger="upload",
                top_analysis_prompt_hint="重点看爆量共性",
            )
            first_jobs = list_analysis_jobs(db_path, batch_id=upload_result.batch_id)

            self.assertGreater(len(first_jobs), 0)
            self.assertEqual(set(first_jobs["trigger"]), {"upload"})
            self.assertEqual(set(first_jobs["prompt_hint"]), {"重点看爆量共性"})

            run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-30",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=db_path,
                env_path=tmp_path / "missing.env",
                enqueue_background_analysis=False,
                background_trigger="page_refresh",
            )
            second_jobs = list_analysis_jobs(db_path, batch_id=upload_result.batch_id)

            self.assertEqual(len(second_jobs), len(first_jobs))

    def test_archived_workflow_emits_progress_messages(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)
            messages: list[str] = []

            run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=tmp_path / "workflow.sqlite3",
                env_path=tmp_path / "missing.env",
                progress_callback=messages.append,
            )

            self.assertIn("正在整理清洗产物", messages)
            self.assertIn("正在读取渠道数据并标准化", messages)
            self.assertIn("正在校验字段完整性与内容类型", messages)
            self.assertIn("正在写入周期库", messages)
            self.assertEqual(messages[-1], "页面数据生成完成")

    def test_archived_workflow_adds_new_columns_to_existing_sqlite_tables(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)
            init_db(db_path)
            with closing(sqlite3.connect(db_path)) as conn:
                conn.execute(
                    "create table canonical_items (batch_id text, platform text, channel text, content_id text)"
                )
                conn.commit()

            result = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            with closing(sqlite3.connect(db_path)) as conn:
                columns = [row[1] for row in conn.execute("pragma table_info(canonical_items)").fetchall()]
                item_count = conn.execute("select count(*) from canonical_items").fetchone()[0]
            self.assertIn("manual_category", columns)
            self.assertIn("ai_category", columns)
            self.assertEqual(item_count, len(result.canonical))

    def test_archived_workflow_reads_previous_batch_for_channel_comparison(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            archive_root = tmp_path / "archive"
            output_root = tmp_path / "outputs"

            first_raw = tmp_path / "first"
            first_raw.mkdir()
            _write_raw_fixture(first_raw)
            run_archived_workflow(
                first_raw,
                "2026-04-01",
                "2026-04-07",
                output_root=output_root,
                archive_root=archive_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            second_raw = tmp_path / "second"
            second_raw.mkdir()
            _write_raw_fixture(second_raw)
            result = run_archived_workflow(
                second_raw,
                "2026-04-08",
                "2026-04-14",
                output_root=output_root,
                archive_root=archive_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            comparison = result.channel_comparison.set_index("channel")
            self.assertIn("总计", comparison.index)
            self.assertIn("spend_change_rate", comparison.columns)
            self.assertEqual(result.comparison_note, "")

            with closing(sqlite3.connect(db_path)) as conn:
                batch_count = conn.execute("select count(*) from upload_batches").fetchone()[0]
            self.assertEqual(batch_count, 2)

    def test_archived_workflow_uses_period_order_instead_of_import_order(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            archive_root = tmp_path / "archive"
            output_root = tmp_path / "outputs"

            april_raw = tmp_path / "april"
            april_raw.mkdir()
            _write_raw_fixture(april_raw)
            april_result = run_archived_workflow(
                april_raw,
                "2026-04-01",
                "2026-04-30",
                output_root=output_root,
                archive_root=archive_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )
            self.assertIn("无历史对比数据", april_result.comparison_note)

            march_raw = tmp_path / "march"
            march_raw.mkdir()
            _write_raw_fixture(march_raw)
            march_result = run_archived_workflow(
                march_raw,
                "2026-03-01",
                "2026-03-31",
                output_root=output_root,
                archive_root=archive_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )
            self.assertIn("无历史对比数据", march_result.comparison_note)

            may_raw = tmp_path / "may"
            may_raw.mkdir()
            _write_raw_fixture(may_raw)
            may_result = run_archived_workflow(
                may_raw,
                "2026-05-01",
                "2026-05-31",
                output_root=output_root,
                archive_root=archive_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            self.assertEqual(may_result.comparison_note, "")
            comparison = may_result.channel_comparison.set_index("channel")
            self.assertIn("总计", comparison.index)
            with closing(sqlite3.connect(db_path)) as conn:
                previous = conn.execute(
                    "select comparison_batch_id from upload_batches where batch_id = ?",
                    (may_result.batch_id,),
                ).fetchone()
            self.assertEqual(previous[0], april_result.batch_id)

    def test_purge_history_state_clears_persisted_results_and_batch_artifacts(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            processed_root = tmp_path / "processed"
            output_root = tmp_path / "outputs"
            first_raw = tmp_path / "first"
            first_raw.mkdir()
            _write_raw_fixture(first_raw)

            run_archived_workflow(
                first_raw,
                "2026-04-01",
                "2026-04-07",
                output_root=output_root,
                processed_root=processed_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )
            second_raw = tmp_path / "second"
            second_raw.mkdir()
            _write_raw_fixture(second_raw)
            run_archived_workflow(
                second_raw,
                "2026-04-08",
                "2026-04-14",
                output_root=output_root,
                processed_root=processed_root,
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            purge_history_state(db_path, output_root, processed_root)

            with closing(sqlite3.connect(db_path)) as conn:
                for table in ["upload_batches", "uploaded_files", "canonical_items", "ai_reports", "channel_comparison_items"]:
                    count = conn.execute(f"select count(*) from {table}").fetchone()[0]
                    self.assertEqual(count, 0, table)
                category_mapping_count = conn.execute("select count(*) from category_mappings").fetchone()[0]
                self.assertEqual(category_mapping_count, 0)

            self.assertEqual(list(output_root.iterdir()), [])
            self.assertEqual(list(processed_root.iterdir()), [])

    def test_archived_workflow_does_not_auto_generate_topic_or_ai_reports(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_path = tmp_path / "workflow.sqlite3"
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-07",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=db_path,
                env_path=tmp_path / "missing.env",
            )

            with closing(sqlite3.connect(db_path)) as conn:
                topic_count = conn.execute(
                    "select count(*) from topic_label_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                ai_count = conn.execute(
                    "select count(*) from ai_reports where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
            self.assertEqual(topic_count, 0)
            self.assertEqual(ai_count, 0)

    def test_workflow_builds_account_audit_and_top_content_items(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_archived_workflow(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                output_root=tmp_path / "outputs",
                archive_root=tmp_path / "archive",
                db_path=tmp_path / "workflow.sqlite3",
                env_path=tmp_path / "missing.env",
            )

            workbook = load_workbook(result.core_recap_xlsx, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["清洗后素材表", "内容复盘表", "不可分析汇总", "匹配覆盖率", "已匹配账号类型分析", "未匹配归因"],
            )
            workbook.close()

            self.assertEqual(len(result.cleaned_asset_table), len(result.canonical))
            self.assertEqual(len(result.top_content_items), 6)
            self.assertEqual(set(result.top_content_items["channel"]), {"B站市场部", "小红书商业化", "抖音商业化", "抖音市场部"})


if __name__ == "__main__":
    unittest.main()
