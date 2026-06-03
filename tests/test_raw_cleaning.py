from pathlib import Path
import shutil
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook
import pandas as pd

from ops_data_workflow.periods import PERIOD_LEVEL_MONTH, PERIOD_LEVEL_WEEK, review_period_from_dates
from ops_data_workflow.raw_cleaning import (
    _additive_metric_columns,
    clean_source_directory,
    clean_raw_period_dir,
    load_cleaned_canonical,
    reset_runtime_data,
)
from ops_data_workflow.storage import init_db
from ops_data_workflow.workflow import run_archived_workflow


def _write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def _bilibili_row(content_id: str, title: str, spend: float = 10.0) -> dict:
    return {
        "视频AVID": content_id.replace("BV", "av"),
        "视频BVID": content_id,
        "视频标题": title,
        "Up主mid": "1622777305",
        "日期": "2026-04-10",
        "花费": spend,
        "展示量": 100,
        "点击量": 10,
        "应用激活数": 2,
        "应用内首次付费次数": 1,
    }


def _xiaohongshu_row(content_id: str, title: str, spend: float = 10.0) -> dict:
    return {
        "时间": "2026-03-02",
        "标题": title,
        "笔记ID": content_id,
        "发布作者": "同花顺投资",
        "类型": "图文",
        "内容分类": "热点行情",
        "笔记链接": f"https://xhs.example/{content_id}",
        "消费": spend,
        "展现量": 100,
        "点击量": 10,
        "激活数": 2,
        "首次付费次数": 1,
    }


class RawCleaningTests(unittest.TestCase):
    def test_additive_metric_detection_uses_configured_columns_only(self):
        frame = pd.DataFrame(
            [
                {
                    "标题": "一条内容",
                    "求和项:总花费": 10,
                    "求和项:未配置指标": 99,
                }
            ]
        )

        self.assertEqual(_additive_metric_columns(frame), ["求和项:总花费"])

    def test_clean_source_directory_prefers_matching_sheet_and_records_ignored_wide_sheet(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "3月总数据" / "0301-0323 小红书数据.xlsx",
                {
                    "1-3月": pd.DataFrame([_xiaohongshu_row("wide-1", "宽周期数据")]),
                    "3.1-3.23": pd.DataFrame([_xiaohongshu_row("march-1", "三月数据")]),
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            self.assertEqual(len(buckets), 1)
            bucket = buckets[0]
            self.assertEqual(bucket.review_period.period_level, PERIOD_LEVEL_MONTH)
            self.assertEqual(bucket.review_period.period_key, "2026-03")
            self.assertEqual(bucket.review_period.data_end, "2026-03-23")
            self.assertTrue(bucket.cleaned_workbook.exists())
            cleaned = load_cleaned_canonical(bucket.cleaned_workbook)
            self.assertEqual(list(cleaned["content_id"]), ["march-1"])
            self.assertEqual(list(cleaned["source_sheet"]), ["3.1-3.23"])

            ignored = pd.read_excel(bucket.cleaned_workbook, sheet_name="忽略sheet")
            self.assertEqual(list(ignored["sheet_name"]), ["1-3月"])
            self.assertIn("宽周期", ignored.iloc[0]["reason"])

    def test_clean_source_directory_merges_multi_account_sheets_and_marks_title_conflicts(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "4月单周数据" / "0410-0416数据" / "B站报表（商业化）.xlsx",
                {
                    "新媒体": pd.DataFrame([_bilibili_row("BV001", "同一标题")]),
                    "研习社": pd.DataFrame([_bilibili_row("BV002", "同一标题", spend=20.0)]),
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(set(cleaned["content_id"]), {"BV001", "BV002"})
            self.assertEqual(set(cleaned["source_sheet"]), {"新媒体", "研习社"})
            self.assertTrue(cleaned["needs_manual_review"].astype(bool).all())
            self.assertTrue(cleaned["review_reasons"].astype(str).str.contains("标题重复但ID不同").all())
            conflicts = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="冲突项")
            self.assertIn("标题重复但ID不同", set(conflicts["issue_type"]))

    def test_clean_source_directory_marks_exact_duplicate_files_without_double_counting(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            frame = pd.DataFrame([_bilibili_row("BV001", "重复文件内容")])
            original = source / "0508-0514 数据" / "B站数据.xlsx"
            duplicate = source / "0508-0514 数据" / "B站数据-副本.xlsx"
            _write_xlsx(original, {"Sheet1": frame})
            shutil.copy2(original, duplicate)

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            duplicate_files = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="重复文件")
            self.assertEqual(len(cleaned), 1)
            self.assertEqual(len(duplicate_files), 1)
            self.assertIn("B站数据-副本.xlsx", duplicate_files.iloc[0]["duplicate_file"])

    def test_clean_source_directory_imports_sum_prefixed_bilibili_and_excludes_total_row(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            frame = pd.DataFrame(
                [
                    {
                        "视频bvid": "BV001",
                        "求和项:总花费": 10.0,
                        "求和项:展示量": 100,
                        "求和项:点击量": 10,
                        "求和项:应用激活数": 2,
                        "求和项:应用内首次付费次数": 1,
                    },
                    {
                        "视频bvid": "BV002",
                        "求和项:总花费": 20.0,
                        "求和项:展示量": 200,
                        "求和项:点击量": 20,
                        "求和项:应用激活数": 4,
                        "求和项:应用内首次付费次数": 2,
                    },
                    {
                        "视频bvid": "",
                        "求和项:总花费": 30.0,
                        "求和项:展示量": 300,
                        "求和项:点击量": 30,
                        "求和项:应用激活数": 6,
                        "求和项:应用内首次付费次数": 3,
                    },
                ]
            )
            _write_xlsx(source / "0508-0514 数据" / "B站数据.xlsx", {"Sheet1": frame})

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(set(cleaned["content_id"]), {"BV001", "BV002"})
            self.assertEqual(float(cleaned["spend"].sum()), 30.0)
            ignored = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="忽略sheet")
            self.assertTrue(ignored["reason"].astype(str).str.contains("汇总行").any())

    def test_clean_raw_period_dir_can_enrich_public_metadata(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            raw_dir = root / "raw"
            _write_xlsx(
                raw_dir / "0508-0514 数据" / "B站数据.xlsx",
                {
                    "Sheet1": pd.DataFrame(
                        [
                            {
                                "视频BVID": "BV1metadata1",
                                "花费": 10.0,
                                "展示量": 100,
                                "点击量": 10,
                                "应用激活数": 2,
                                "应用内首次付费次数": 1,
                            }
                        ]
                    )
                },
            )
            period = review_period_from_dates(
                pd.Timestamp("2026-05-08").date(),
                pd.Timestamp("2026-05-14").date(),
                PERIOD_LEVEL_WEEK,
            )

            bucket = clean_raw_period_dir(
                raw_dir,
                period,
                default_year=2026,
                output_dir=root / "processed",
                metadata_enrichment_mode="safe_public",
                metadata_cache_dir=root / "metadata-cache",
                fetch_bilibili_metadata=lambda bvid: {
                    "id": bvid,
                    "link": f"https://www.bilibili.com/video/{bvid}/",
                    "title": "补全标题",
                    "tags": "财经",
                    "published_at": "2026-05-09",
                },
            )

            cleaned = load_cleaned_canonical(bucket.cleaned_workbook)
            row = cleaned.iloc[0]
            self.assertEqual(row["content_url"], "https://www.bilibili.com/video/BV1metadata1/")
            self.assertEqual(row["title"], "补全标题")
            self.assertEqual(row["source_time"], "2026-05-09")
            self.assertEqual(row["metadata_tags"], "财经")
            self.assertEqual(row["metadata_source"], "bilibili_public_api")

            manifest = pd.read_json(bucket.manifest_path, typ="series").to_dict()
            self.assertEqual(manifest["metadata_enrichment"]["mode"], "safe_public")
            self.assertEqual(manifest["metadata_enrichment"]["filled_rows"], 1)

    def test_clean_source_directory_dedupes_bilibili_duplicate_summary_sheets_by_unit_name(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            sheet2 = pd.DataFrame(
                [
                    {
                        "单元名称": "BV001",
                        "求和项:总花费": 10.0,
                        "求和项:应用激活数": 2,
                        "求和项:应用内首次付费次数": 1,
                    },
                    {
                        "单元名称": "BV002",
                        "求和项:总花费": 20.0,
                        "求和项:应用激活数": 4,
                        "求和项:应用内首次付费次数": 2,
                    },
                    {
                        "单元名称": "BV000",
                        "求和项:总花费": 0.0,
                        "求和项:应用激活数": 0,
                        "求和项:应用内首次付费次数": 0,
                    },
                ]
            )
            sheet1 = pd.DataFrame(
                [
                    {
                        "单元名称": "BV001",
                        "总花费": 10.0,
                        "应用激活数": 2,
                        "应用内首次付费次数": 1,
                        "激活成本": 5.0,
                        "付费成本": 10.0,
                    },
                    {
                        "单元名称": "BV002",
                        "总花费": 20.0,
                        "应用激活数": 4,
                        "应用内首次付费次数": 2,
                        "激活成本": 5.0,
                        "付费成本": 10.0,
                    },
                    {
                        "单元名称": "BV000",
                        "总花费": 0.0,
                        "应用激活数": 0,
                        "应用内首次付费次数": 0,
                        "激活成本": "",
                        "付费成本": "",
                    },
                ]
            )
            _write_xlsx(source / "0515-0521 数据" / "B站.xlsx", {"Sheet2": sheet2, "Sheet1": sheet1})

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(set(cleaned["content_id"]), {"BV001", "BV002"})
            self.assertEqual(len(cleaned), 2)
            self.assertEqual(float(cleaned["spend"].sum()), 30.0)
            self.assertEqual(float(cleaned["activations"].sum()), 6.0)
            self.assertEqual(float(cleaned["first_pay_count"].sum()), 3.0)

    def test_clean_source_directory_records_every_workbook_sheet_as_imported_or_ignored(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0515-0521 数据" / "抖音商业化.xlsx",
                {
                    "Sheet1": pd.DataFrame([{"说明": "不是投放明细"}]),
                    "Sheet2": pd.DataFrame(
                        [
                            {
                                "视频标题": "有效投放内容",
                                "视频id": "dy-1",
                                "素材ID": "mat-1",
                                "消耗": 10.0,
                                "展示数": 100,
                                "激活数": 2,
                                "付费次数": 1,
                            }
                        ]
                    ),
                    "空白": pd.DataFrame(),
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            workbook_path = source / "0515-0521 数据" / "抖音商业化.xlsx"
            expected_sheets = set(pd.ExcelFile(workbook_path).sheet_names)
            import_log = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="导入日志")
            ignored = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="忽略sheet")
            recorded_sheets = set(import_log["sheet_name"].dropna().astype(str)) | set(
                ignored["sheet_name"].dropna().astype(str)
            )

            self.assertEqual(recorded_sheets, expected_sheets)

    def test_clean_source_directory_fills_douyin_grouped_image_text_content_type(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            workbook_path = source / "0527-0602 数据" / "抖音市场部数据.xlsx"
            _write_xlsx(
                workbook_path,
                {
                    "Sheet2": pd.DataFrame(
                        [
                            {
                                "创建时间": "图文",
                                "时长": "",
                                "视频标题": "存多少钱才能提前退休",
                                "视频链接": "",
                                "消耗": 93768.8,
                                "展示数": 7736899,
                                "激活数": 2706,
                                "付费次数": 910,
                            },
                            {
                                "创建时间": "",
                                "时长": "",
                                "视频标题": "如何字字不提股票，但一听就是炒股人",
                                "视频链接": "",
                                "消耗": 15095.017,
                                "展示数": 1642653,
                                "激活数": 403,
                                "付费次数": 165,
                            },
                            {
                                "创建时间": "",
                                "时长": "",
                                "视频标题": "全国在校大学生招募",
                                "视频链接": "",
                                "消耗": 12967.43,
                                "展示数": 679457,
                                "激活数": 335,
                                "付费次数": 111,
                            },
                            {
                                "创建时间": "2026-05-30 10:00:00",
                                "时长": "00:15",
                                "视频标题": "普通视频内容",
                                "视频链接": "",
                                "消耗": 10,
                                "展示数": 100,
                                "激活数": 1,
                                "付费次数": 1,
                            },
                        ]
                    )
                },
            )
            workbook = load_workbook(workbook_path)
            worksheet = workbook["Sheet2"]
            worksheet.merge_cells("A2:B4")
            workbook.save(workbook_path)
            workbook.close()

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            image_text = cleaned[cleaned["title"].isin(
                [
                    "存多少钱才能提前退休",
                    "如何字字不提股票，但一听就是炒股人",
                    "全国在校大学生招募",
                ]
            )]
            self.assertEqual(len(image_text), 3)
            self.assertEqual(set(image_text["manual_category"]), {"图文"})
            self.assertAlmostEqual(float(image_text["spend"].sum()), 121831.247)
            video = cleaned[cleaned["title"].eq("普通视频内容")].iloc[0]
            self.assertTrue(pd.isna(video["manual_category"]) or video["manual_category"] == "")

    def test_clean_source_directory_sums_xiaohongshu_and_bilibili_id_duplicates_without_conflicts(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0508-0514 数据" / "小红书商业化.xlsx",
                {
                    "kos账号笔记投放数据": pd.DataFrame(
                        [
                            _xiaohongshu_row("note-dup", "5.7日段永平调仓买入泡泡玛特，头像也换了！", 10.0),
                            _xiaohongshu_row(
                                "note-dup",
                                "5.7日段永平调仓买入泡泡玛特，头像也换了！ #同花顺APP #同花顺资讯",
                                20.0,
                            ),
                        ]
                    )
                },
            )
            _write_xlsx(
                source / "0508-0514 数据" / "B站数据.xlsx",
                {
                    "Sheet1": pd.DataFrame(
                        [
                            _bilibili_row("BV001", "同一条B站内容", 30.0),
                            _bilibili_row("BV001", "同一条B站内容", 30.0),
                        ]
                    )
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(len(cleaned), 2)
            xhs = cleaned[cleaned["channel"].eq("小红书商业化")].iloc[0]
            bilibili = cleaned[cleaned["channel"].eq("B站")].iloc[0]
            self.assertEqual(xhs["title"], "5.7日段永平调仓买入泡泡玛特，头像也换了！ #同花顺APP #同花顺资讯")
            self.assertEqual(float(xhs["spend"]), 30.0)
            self.assertEqual(float(xhs["impressions"]), 200.0)
            self.assertEqual(float(xhs["activations"]), 4.0)
            self.assertEqual(float(bilibili["spend"]), 60.0)
            self.assertEqual(float(bilibili["impressions"]), 200.0)
            self.assertEqual(float(bilibili["activations"]), 4.0)

            duplicate_content = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="重复内容")
            self.assertEqual(len(duplicate_content), 2)
            self.assertEqual(set(duplicate_content["merged_row_count"]), {2})
            conflicts = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="冲突项")
            self.assertTrue(conflicts.empty)

    def test_clean_source_directory_sums_bilibili_impression_aliases(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0508-0514 数据" / "B站数据.xlsx",
                {
                    "Sheet1": pd.DataFrame(
                        [
                            {
                                "视频AVID": "av001",
                                "视频BVID": "BV001",
                                "视频标题": "B站展示量别名内容",
                                "花费": 30.0,
                                "视频展示量": 111,
                                "曝光转化率": 0.12,
                                "千次展示费用": 20.0,
                                "点击量": 10,
                                "应用激活数": 2,
                                "应用内首次付费次数": 1,
                            },
                            {
                                "视频AVID": "av001",
                                "视频BVID": "BV001",
                                "视频标题": "B站展示量别名内容",
                                "花费": 40.0,
                                "曝光量(次)": 222,
                                "点击量": 20,
                                "应用激活数": 3,
                                "应用内首次付费次数": 2,
                            },
                        ]
                    )
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            bilibili = cleaned[cleaned["channel"].eq("B站")].iloc[0]
            self.assertEqual(float(bilibili["impressions"]), 333.0)
            self.assertEqual(float(bilibili["spend"]), 70.0)
            self.assertEqual(float(bilibili["activations"]), 5.0)
            self.assertEqual(float(bilibili["first_pay_count"]), 3.0)

    def test_clean_source_directory_applies_ledger_and_writes_channel_clean_workbook(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            period_dir = source / "0508-0514 数据"
            _write_xlsx(
                period_dir / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 12",
                                "内容链接": "1.28 tRk:/ 人和人的缘分就像炒股 # 同花顺股友说 # 投资 https://v.douyin.com/abc/ 复制此链接，打开抖音搜索，直接观看视频！",
                                "账号": "投资号",
                                "内容类型": "股友说",
                            }
                        ]
                    )
                },
            )
            _write_xlsx(
                period_dir / "抖音商业化.xlsx",
                {
                    "Sheet2": pd.DataFrame(
                        [
                            {
                                "视频标题": "人和人的缘分就像炒股 #投资",
                                "消耗": 90,
                                "展示数": 9000,
                                "激活数": 9,
                                "付费次数": 3,
                            }
                        ]
                    )
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            bucket = buckets[0]
            cleaned = load_cleaned_canonical(bucket.cleaned_workbook)
            row = cleaned.iloc[0]
            self.assertEqual(row["account"], "投资号")
            self.assertEqual(row["manual_category"], "股友说")
            self.assertEqual(row["content_url"], "https://v.douyin.com/abc/")
            self.assertEqual(row["ledger_match_source"], "唯一标题")
            self.assertEqual(row["manual_category_source"], "投稿台账补全")

            channel_clean = bucket.raw_dir / "channel_clean" / "抖音商业化_clean.xlsx"
            self.assertTrue(channel_clean.exists())
            display = pd.read_excel(channel_clean, sheet_name="清理后明细")
            self.assertEqual(
                list(display.columns),
                [
                    "周期",
                    "渠道",
                    "账号",
                    "内容形式",
                    "内容类型",
                    "内容分类",
                    "标题",
                    "id/BV或者唯一标识",
                    "内容链接",
                    "消耗",
                    "曝光量",
                    "激活数",
                    "激活成本",
                    "付费",
                    "付费成本",
                    "匹配来源",
                    "复核原因",
                ],
            )
            display_row = display.iloc[0]
            self.assertEqual(display_row["渠道"], "抖音商业化")
            self.assertEqual(display_row["账号"], "投资号")
            self.assertEqual(display_row["内容形式"], "视频")
            self.assertEqual(display_row["内容类型"], "股友说")
            self.assertEqual(display_row["内容分类"], "股友说")
            self.assertEqual(display_row["id/BV或者唯一标识"], "人和人的缘分就像炒股")
            self.assertEqual(display_row["内容链接"], "https://v.douyin.com/abc/")
            self.assertEqual(display_row["匹配来源"], "唯一标题")
            self.assertTrue(pd.isna(display_row["复核原因"]) or display_row["复核原因"] == "")

    def test_clean_source_directory_generates_row_ids_for_identityless_social_details_and_excludes_total(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            frame = pd.DataFrame(
                [
                    {"创意名称": "", "花费": 10.0, "曝光次数": 100, "点击次数": 10, "APP激活次数": 2, "注册次数": 1},
                    {"创意名称": "", "花费": 20.0, "曝光次数": 200, "点击次数": 20, "APP激活次数": 4, "注册次数": 2},
                    {"创意名称": "", "花费": 30.0, "曝光次数": 300, "点击次数": 30, "APP激活次数": 6, "注册次数": 3},
                ]
            )
            _write_xlsx(source / "0508-0514 数据" / "微信市场部.xlsx", {"Sheet1": frame})

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(len(cleaned), 2)
            self.assertEqual(set(cleaned["channel"]), {"微信市场部"})
            self.assertEqual(set(cleaned["platform"]), {"微信"})
            self.assertEqual(set(cleaned["platform_group"]), {"微信"})
            self.assertEqual(float(cleaned["spend"].sum()), 30.0)
            self.assertTrue(cleaned["content_id"].astype(str).str.startswith("row:").all())
            ignored = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="忽略sheet")
            self.assertTrue(ignored["reason"].astype(str).str.contains("汇总行").any())

    def test_clean_source_directory_excludes_group_subtotal_rows(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            frame = pd.DataFrame(
                [
                    {
                        "素材ID": "dy-1",
                        "视频标题": "达人内容一",
                        "视频链接": "https://douyin.example/1",
                        "消耗": 10.0,
                        "展示数": 100,
                        "激活数": 2,
                        "付费次数": 1,
                        "内容类型": "达人内容",
                    },
                    {
                        "素材ID": "dy-2",
                        "视频标题": "股友说一",
                        "视频链接": "https://douyin.example/2",
                        "消耗": 20.0,
                        "展示数": 200,
                        "激活数": 4,
                        "付费次数": 2,
                        "内容类型": "股友说",
                    },
                    {
                        "素材ID": "",
                        "视频标题": "",
                        "视频链接": "",
                        "消耗": 30.0,
                        "展示数": 300,
                        "激活数": 6,
                        "付费次数": 3,
                        "内容类型": "",
                    },
                    {
                        "素材ID": "",
                        "视频标题": "",
                        "视频链接": "",
                        "消耗": 10.0,
                        "展示数": 100,
                        "激活数": 2,
                        "付费次数": 1,
                        "内容类型": "",
                    },
                ]
            )
            _write_xlsx(source / "0508-0514 数据" / "抖音市场部数据.xlsx", {"Sheet2": frame})

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(set(cleaned["content_id"]), {"https://douyin.example/1", "https://douyin.example/2"})
            self.assertEqual(float(cleaned["spend"].sum()), 30.0)
            ignored = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="忽略sheet")
            reasons = "\n".join(ignored["reason"].astype(str).tolist())
            self.assertIn("汇总行", reasons)
            self.assertIn("分组小计行", reasons)

    def test_clean_source_directory_ignores_rows_without_statistical_metrics(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0508-0514 数据" / "抖音商业化.xlsx",
                {
                    "Sheet2": pd.DataFrame(
                        [
                            {
                                "视频标题": "只有标题没有统计指标",
                                "视频链接": "https://douyin.example/no-metric",
                                "消耗": "",
                                "展示数": "",
                                "激活数": "",
                                "付费次数": "",
                            },
                            {
                                "视频标题": "有效投放内容",
                                "视频链接": "https://douyin.example/valid",
                                "消耗": 10,
                                "展示数": 100,
                                "激活数": 2,
                                "付费次数": 1,
                            }
                        ]
                    ),
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(list(cleaned["title"]), ["有效投放内容"])
            self.assertEqual(float(cleaned["spend"].sum()), 10.0)

    def test_clean_source_directory_keeps_identified_exposure_only_content(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0410-0416 数据" / "抖音原生内容（商业化）.xlsx",
                {
                    "Sheet1": pd.DataFrame(
                        [
                            {
                                "视频标题": "有标题的自然流量内容",
                                "视频链接": "https://douyin.example/organic",
                                "内容类型": "股友说",
                                "展示数": 5317,
                            }
                        ]
                    )
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(list(cleaned["content_id"]), ["https://douyin.example/organic"])
            self.assertEqual(float(cleaned["impressions"].sum()), 5317.0)
            self.assertTrue(cleaned["spend"].isna().all())

    def test_clean_source_directory_ignores_identityless_single_metric_artifacts(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0410-0416 数据" / "抖音原生-达人数据情况（商业化）.xlsx",
                {
                    "Sheet4": pd.DataFrame(
                        [
                            {"视频链接": "", "消耗": "", "展示数": 5317, "激活数": "", "付费数": ""},
                            {
                                "视频链接": "",
                                "消耗": 10,
                                "展示数": 100,
                                "激活数": 2,
                                "付费数": 1,
                            },
                        ]
                    )
                },
            )

            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")

            cleaned = load_cleaned_canonical(buckets[0].cleaned_workbook)
            self.assertEqual(len(cleaned), 1)
            self.assertTrue(cleaned["content_id"].astype(str).str.startswith("row:").all())
            self.assertEqual(float(cleaned["spend"].sum()), 10.0)
            ignored = pd.read_excel(buckets[0].cleaned_workbook, sheet_name="忽略sheet")
            self.assertTrue(ignored["reason"].astype(str).str.contains("无可追溯标识").any())

    def test_archived_workflow_uses_cleaned_workbook_instead_of_noisy_raw_files(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source"
            _write_xlsx(
                source / "0508-0514 数据" / "B站数据.xlsx",
                {"Sheet1": pd.DataFrame([_bilibili_row("BV001", "清洗后内容")])},
            )
            buckets = clean_source_directory(source, root / "data" / "raw", default_year=2026, import_id="import-test")
            raw_dir = buckets[0].raw_dir
            _write_xlsx(
                raw_dir / "噪声人工统计.xlsx",
                {"Sheet1": pd.DataFrame([{"渠道": "B站", "消耗": 999999, "激活": 999999}])},
            )

            result = run_archived_workflow(
                raw_dir,
                "",
                "",
                output_root=root / "outputs",
                archive_root=root / "archive",
                db_path=root / "data" / "workflow.sqlite3",
                env_path=root / "missing.env",
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(list(result.canonical["content_id"]), ["BV001"])
            self.assertNotIn("噪声人工统计.xlsx", set(result.canonical["source_file"]))

    def test_archived_workflow_backfills_core_metrics_from_cleaned_raw_columns(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            raw_dir = root / "data" / "raw" / "20260515-20260521"
            raw_dir.mkdir(parents=True)
            _write_xlsx(
                raw_dir / "cleaned.xlsx",
                {
                    "清洗后明细": pd.DataFrame(
                        [
                            {
                                "platform": "B站",
                                "platform_group": "B站",
                                "channel": "B站",
                                "content_id": "BV-cleaned",
                                "material_id": "BV-cleaned",
                                "account_id": "1622777305",
                                "title": "清洗后B站内容",
                                "manual_category": "产品科普",
                                "spend": pd.NA,
                                "impressions": pd.NA,
                                "clicks": pd.NA,
                                "activations": pd.NA,
                                "first_pay_count": pd.NA,
                                "source_file": "B站.xlsx",
                                "source_sheet": "Sheet1",
                                "source_row": 2,
                                "raw__B站__总花费": 100.0,
                                "raw__B站__视频展示量": 4321,
                                "raw__B站__曝光转化率": 0.12,
                                "raw__B站__千次展示费用": 23.5,
                                "raw__B站__点击量": 321,
                                "raw__B站__APP激活数": 11,
                                "raw__B站__激活成本": 9.1,
                                "raw__B站__首次付费数": 4,
                                "raw__B站__付费成本": 25.0,
                            }
                        ]
                    )
                },
            )
            (raw_dir / "period_manifest.json").write_text(
                '{"files":["cleaned.xlsx"],"cleaned_workbook":"cleaned.xlsx"}',
                encoding="utf-8",
            )

            result = run_archived_workflow(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                output_root=root / "outputs",
                archive_root=root / "archive",
                db_path=root / "data" / "workflow.sqlite3",
                env_path=root / "missing.env",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = result.canonical.iloc[0]
            self.assertAlmostEqual(row["spend"], 100.0)
            self.assertAlmostEqual(row["impressions"], 4321.0)
            self.assertAlmostEqual(row["clicks"], 321.0)
            self.assertAlmostEqual(row["activations"], 11.0)
            self.assertAlmostEqual(row["first_pay_count"], 4.0)
            channel_summary = result.channel_summary.set_index("channel")
            self.assertAlmostEqual(channel_summary.loc["B站", "spend"], 100.0)
            self.assertAlmostEqual(channel_summary.loc["B站", "impressions"], 4321.0)
            self.assertAlmostEqual(channel_summary.loc["B站", "activations"], 11.0)
            self.assertAlmostEqual(channel_summary.loc["B站", "first_pay_count"], 4.0)

    def test_reset_runtime_data_clears_generated_runtime_dirs_and_reinitializes_db(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            for relative in [
                "data/raw/20260508-20260514/old.xlsx",
                "data/file_backup/old/raw/old.xlsx",
                "archive/old/raw/old.xlsx",
                "processed/old/cleaned.xlsx",
                "outputs/old/report.html",
                "output/playwright/old.png",
            ]:
                path = root / relative
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("old", encoding="utf-8")
            init_db(root / ".runtime" / "workflow.sqlite3")

            reset_runtime_data(root)

            self.assertTrue((root / ".runtime" / "workflow.sqlite3").exists())
            self.assertTrue((root / "data" / "reference").exists())
            self.assertTrue((root / "data" / "months").exists())
            self.assertTrue((root / "data" / "weeks").exists())
            self.assertTrue((root / "data" / "raw" / "20260508-20260514" / "old.xlsx").exists())
            self.assertFalse((root / "data" / "file_backup" / "old").exists())
            self.assertFalse((root / "archive" / "old").exists())
            self.assertFalse((root / "processed" / "old").exists())
            self.assertFalse((root / "outputs" / "old").exists())
            self.assertFalse((root / "output" / "playwright" / "old.png").exists())


if __name__ == "__main__":
    unittest.main()
