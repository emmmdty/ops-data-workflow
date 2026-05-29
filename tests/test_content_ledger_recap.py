from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

import pandas as pd
from openpyxl import load_workbook

from ops_data_workflow.content_ledger import apply_content_ledger, load_content_ledger
from ops_data_workflow.pipeline import analyze_canonical_frame, analyze_input_dir
from ops_data_workflow.recap import build_recap_summary
from ops_data_workflow.reporting import write_outputs


def _write_xlsx(path: Path, sheets: dict[str, pd.DataFrame], *, startrow: int = 0) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)


class ContentLedgerTests(unittest.TestCase):
    def test_harvester_ledger_rows_normalize_current_and_legacy_sheet_shapes(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            ledger_path = root / "原生内容投稿.xlsx"
            _write_xlsx(
                ledger_path,
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "1.28 tRk:/ 人和人的缘分就像炒股 # 同花顺股友说 # 投资 https://v.douyin.com/abc/ 复制此链接，打开Dou音搜索，直接观看视频！",
                                "标题": "",
                                "tag词": "",
                                "筛选状态": "通过",
                                "账号": "投资号",
                                "内容类型": "",
                                "内容类型标签审核": "",
                            }
                        ]
                    ),
                    "小红书渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "https://www.xiaohongshu.com/discovery/item/6a115b0e00000000360031a6?source=webshare",
                                "笔记ID": "",
                                "账号": "投资号",
                                "内容类型": "资讯",
                            }
                        ]
                    ),
                    "B站渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "https://www.bilibili.com/video/BV1crLm6yE5V/",
                                "短链id": "",
                                "账号": "投资号",
                            }
                        ]
                    ),
                },
            )

            ledger = load_content_ledger(root, default_year=2026)

            self.assertEqual(set(ledger["platform"]), {"抖音", "小红书", "B站"})
            douyin = ledger[ledger["platform"].eq("抖音")].iloc[0]
            self.assertEqual(douyin["account"], "投资号")
            self.assertEqual(douyin["title"], "人和人的缘分就像炒股")
            self.assertEqual(douyin["tags"], "#同花顺股友说 #投资")
            self.assertEqual(douyin["content_type"], "股友说")
            xhs = ledger[ledger["platform"].eq("小红书")].iloc[0]
            self.assertEqual(xhs["content_id"], "6a115b0e00000000360031a6")
            bili = ledger[ledger["platform"].eq("B站")].iloc[0]
            self.assertEqual(bili["content_id"], "BV1crLm6yE5V")

    def test_content_ledger_backfills_douyin_category_and_link_by_account_title(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "1.28 tRk:/ 人和人的缘分就像炒股 # 同花顺股友说 # 投资 https://v.douyin.com/abc/ 复制此链接，打开Dou音搜索，直接观看视频！",
                                "账号": "投资号",
                                "内容类型": "",
                            }
                        ]
                    )
                },
            )
            _write_xlsx(
                root / "抖音商业化.xlsx",
                {
                    "Sheet2": pd.DataFrame(
                        [
                            {
                                "视频标题": "人和人的缘分就像炒股",
                                "账号": "同花顺投资",
                                "消耗": 100,
                                "展示数": 1000,
                                "激活数": 10,
                                "付费次数": 2,
                            }
                        ]
                    )
                },
            )

            analysis = analyze_input_dir(
                root,
                "2026-05-19",
                "2026-05-25",
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(len(analysis.canonical), 1)
            row = analysis.canonical.iloc[0]
            self.assertEqual(row["content_category"], "股友说")
            self.assertEqual(row["category_status"], "投稿台账补全")
            self.assertEqual(row["content_url"], "https://v.douyin.com/abc/")
            self.assertEqual(row["ledger_match_source"], "账号+标题")
            self.assertEqual(row["first_pay_cost"], 50)

    def test_xiaohongshu_ledger_backfills_market_rows_by_explore_id_without_account(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "小红书渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "https://www.xiaohongshu.com/explore/6a115b0e00000000360031a6?xsec_source=pc_ad_export",
                                "标题": "小红书投教内容 #同花顺",
                                "账号": "投资号",
                                "内容类型": "投教问答",
                            }
                        ]
                    )
                },
            )
            _write_xlsx(
                root / "小红书市场部.xlsx",
                {
                    "新户付费": pd.DataFrame(
                        [
                            {
                                "笔记/素材ID": "6a115b0e00000000360031a6",
                                "笔记/素材链接": "https://www.xiaohongshu.com/explore/6a115b0e00000000360031a6?xsec_source=pc_ad_export",
                                "账号": "同花顺投资",
                                "消费": 88,
                                "展现量": 8800,
                                "激活数": 8,
                                "首次付费次数": 2,
                            }
                        ]
                    )
                },
            )

            analysis = analyze_input_dir(
                root,
                "2026-05-19",
                "2026-05-25",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = analysis.canonical.iloc[0]
            self.assertEqual(row["title"], "小红书投教内容 #同花顺")
            self.assertEqual(row["account"], "同花顺投资")
            self.assertEqual(row["content_category"], "投教问答")
            self.assertEqual(row["ledger_content_type"], "投教问答")
            self.assertEqual(row["ledger_match_source"], "id")

    def test_douyin_unique_tagless_title_can_fill_missing_account_and_category(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "1.28 tRk:/ 人和人的缘分就像炒股 # 同花顺股友说 # 投资 https://v.douyin.com/abc/ 复制此链接，打开抖音搜索，直接观看视频！",
                                "账号": "投资号",
                                "内容类型": "股友说",
                            }
                        ]
                    )
                },
            )
            ledger = load_content_ledger(root, default_year=2026)
            canonical = pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "人和人的缘分就像炒股 #投资",
                        "account": "",
                        "manual_category": "",
                        "content_url": "",
                    }
                ]
            )

            enriched = apply_content_ledger(canonical, ledger)

            row = enriched.iloc[0]
            self.assertEqual(row["account"], "投资号")
            self.assertEqual(row["manual_category"], "股友说")
            self.assertEqual(row["content_url"], "https://v.douyin.com/abc/")
            self.assertEqual(row["ledger_match_source"], "唯一标题")
            self.assertEqual(row["match_risk_reason"], "")

    def test_douyin_duplicate_tagless_title_requires_review_without_autofill(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "内容链接": "同一个标题 # 投资 https://v.douyin.com/abc/",
                                "账号": "投资号",
                                "内容类型": "股友说",
                            },
                            {
                                "编号": 2,
                                "内容链接": "同一个标题 # 理财 https://v.douyin.com/def/",
                                "账号": "理财",
                                "内容类型": "图文",
                            },
                        ]
                    )
                },
            )
            ledger = load_content_ledger(root, default_year=2026)
            canonical = pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "同一个标题 #未知tag",
                        "account": "",
                        "manual_category": "",
                    }
                ]
            )

            enriched = apply_content_ledger(canonical, ledger)

            row = enriched.iloc[0]
            self.assertEqual(row["account"], "")
            self.assertEqual(row["manual_category"], "")
            self.assertEqual(row["ledger_match_source"], "唯一标题")
            self.assertEqual(row["match_risk_level"], "需复核")
            self.assertIn("投稿台账存在 2 条同标题记录", row["match_risk_reason"])

    def test_douyin_share_prefix_cleaning_allows_unique_title_autofill(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "/05 J@i.pd 离大谱!炒股还能领亏损补贴??? https://v.douyin.com/abc/ 复制此链接，打开抖音搜索，直接观看视频！",
                                "账号": "投资号",
                                "内容类型": "资讯",
                            }
                        ]
                    )
                },
            )
            ledger = load_content_ledger(root, default_year=2026)
            canonical = pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "离大谱!炒股还能领亏损补贴??? #投资",
                        "account": "",
                        "manual_category": "",
                        "content_url": "",
                    }
                ]
            )

            enriched = apply_content_ledger(canonical, ledger)

            row = enriched.iloc[0]
            self.assertEqual(row["account"], "投资号")
            self.assertEqual(row["manual_category"], "资讯")
            self.assertEqual(row["ledger_match_source"], "唯一标题")
            self.assertEqual(row["match_risk_reason"], "")

    def test_douyin_fuzzy_title_marks_review_candidate_without_autofill(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 22",
                                "内容链接": "炒股还能领亏损补贴吗？这件事很多人不知道 https://v.douyin.com/abc/",
                                "账号": "投资号",
                                "内容类型": "资讯",
                            }
                        ]
                    )
                },
            )
            ledger = load_content_ledger(root, default_year=2026)
            canonical = pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "炒股还能领亏损补贴吗？",
                        "account": "",
                        "manual_category": "",
                        "content_url": "",
                    }
                ]
            )

            enriched = apply_content_ledger(canonical, ledger)

            row = enriched.iloc[0]
            self.assertEqual(row["account"], "")
            self.assertEqual(row["manual_category"], "")
            self.assertEqual(row["ledger_match_source"], "模糊标题")
            self.assertEqual(row["ledger_content_type"], "资讯")
            self.assertEqual(row["ledger_source_row"], 2)
            self.assertEqual(row["match_risk_level"], "需复核")
            self.assertIn("标题近似匹配，需确认", row["match_risk_reason"])

    def test_douyin_id_bridge_backfills_before_title_matching(self):
        canonical = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "v02033g10000bridge",
                    "title": "投放侧标题已经完全改写",
                    "account": "",
                    "manual_category": "",
                    "content_url": "",
                }
            ]
        )
        bridge = pd.DataFrame(
            [
                {
                    "id_type": "content_id",
                    "id_value": "v02033g10000bridge",
                    "account": "投资号",
                    "content_type": "股友说",
                    "content_url": "https://v.douyin.com/bridge/",
                    "source_file": "原生内容投稿.xlsx",
                    "source_sheet": "抖音渠道",
                    "source_row": 12,
                    "title_key_no_tags": "原始台账标题",
                }
            ]
        )

        enriched = apply_content_ledger(
            canonical,
            pd.DataFrame(columns=["platform", "content_id", "title_key_no_tags"]),
            douyin_id_bridge=bridge,
        )

        row = enriched.iloc[0]
        self.assertEqual(row["account"], "投资号")
        self.assertEqual(row["manual_category"], "股友说")
        self.assertEqual(row["content_url"], "https://v.douyin.com/bridge/")
        self.assertEqual(row["ledger_match_source"], "反馈ID桥表")
        self.assertEqual(row["ledger_source_row"], 12)

    def test_douyin_id_bridge_conflict_requires_review_without_autofill(self):
        canonical = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "v02033g10000conflict",
                    "title": "投放侧标题",
                    "account": "",
                    "manual_category": "",
                }
            ]
        )
        bridge = pd.DataFrame(
            [
                {
                    "id_type": "content_id",
                    "id_value": "v02033g10000conflict",
                    "account": "投资号",
                    "content_type": "股友说",
                },
                {
                    "id_type": "content_id",
                    "id_value": "v02033g10000conflict",
                    "account": "理财",
                    "content_type": "图文",
                },
            ]
        )

        enriched = apply_content_ledger(
            canonical,
            pd.DataFrame(columns=["platform", "content_id", "title_key_no_tags"]),
            douyin_id_bridge=bridge,
        )

        row = enriched.iloc[0]
        self.assertEqual(row["account"], "")
        self.assertEqual(row["manual_category"], "")
        self.assertEqual(row["ledger_match_source"], "反馈ID桥表")
        self.assertEqual(row["match_risk_level"], "需复核")
        self.assertIn("抖音ID桥表存在 2 条同ID记录", row["match_risk_reason"])

    def test_douyin_duplicate_tagless_title_prefers_earliest_published_date(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "抖音渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "05 23",
                                "内容链接": "同一个标题 # 后发 https://v.douyin.com/later/",
                                "账号": "理财",
                                "内容类型": "图文",
                            },
                            {
                                "编号": 2,
                                "投稿时间": "05 20",
                                "内容链接": "同一个标题 # 先发 https://v.douyin.com/early/",
                                "账号": "投资号",
                                "内容类型": "股友说",
                            },
                        ]
                    )
                },
            )
            ledger = load_content_ledger(root, default_year=2026)
            canonical = pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "同一个标题 #任意tag",
                        "account": "",
                        "manual_category": "",
                        "content_url": "",
                    }
                ]
            )

            enriched = apply_content_ledger(canonical, ledger)

            row = enriched.iloc[0]
            self.assertEqual(row["account"], "投资号")
            self.assertEqual(row["manual_category"], "股友说")
            self.assertEqual(row["content_url"], "https://v.douyin.com/early/")
            self.assertEqual(row["ledger_match_source"], "唯一标题")
            self.assertEqual(row["match_risk_reason"], "")

    def test_xiaohongshu_duplicate_id_prefers_earliest_published_date(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            _write_xlsx(
                root / "原生内容投稿.xlsx",
                {
                    "小红书渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "投稿时间": "2026-05-23",
                                "内容链接": "https://www.xiaohongshu.com/explore/6a115b0e00000000360031a6?xsec_source=later",
                                "账号": "理财",
                                "内容类型": "图文",
                            },
                            {
                                "编号": 2,
                                "投稿时间": "2026-05-20",
                                "内容链接": "https://www.xiaohongshu.com/explore/6a115b0e00000000360031a6?xsec_source=early",
                                "账号": "投资号",
                                "内容类型": "资讯",
                            },
                        ]
                    )
                },
            )
            ledger = load_content_ledger(root, default_year=2026)
            canonical = pd.DataFrame(
                [
                    {
                        "platform": "小红书商业化",
                        "platform_group": "小红书",
                        "channel": "小红书商业化",
                        "content_id": "6a115b0e00000000360031a6",
                        "title": "",
                        "account": "",
                        "manual_category": "",
                        "content_url": "",
                    }
                ]
            )

            enriched = apply_content_ledger(canonical, ledger)

            row = enriched.iloc[0]
            self.assertEqual(row["account"], "投资号")
            self.assertEqual(row["manual_category"], "资讯")
            self.assertEqual(row["content_url"], "https://www.xiaohongshu.com/explore/6a115b0e00000000360031a6?xsec_source=early")
            self.assertEqual(row["ledger_match_source"], "id")
            self.assertEqual(row["match_risk_reason"], "")

    def test_configured_feishu_export_paths_load_as_ledger_sources(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            export_path = root / "feishu_exports" / "原生内容投稿.xlsx"
            _write_xlsx(
                export_path,
                {
                    "小红书渠道": pd.DataFrame(
                        [
                            {
                                "编号": 1,
                                "内容链接": "https://www.xiaohongshu.com/explore/6a115b0e00000000360031a6",
                                "账号": "投资号",
                                "内容类型": "资讯",
                            }
                        ]
                    )
                },
            )
            config_path = root / "config" / "feishu_sources.yml"
            config_path.parent.mkdir(parents=True)
            config_path.write_text(
                "content_ledgers:\n"
                "  - type: feishu_sheet_export\n"
                "    path: ../feishu_exports/原生内容投稿.xlsx\n",
                encoding="utf-8",
            )

            ledger = load_content_ledger(root / "empty_input", default_year=2026, config_path=config_path)

            self.assertEqual(len(ledger), 1)
            row = ledger.iloc[0]
            self.assertEqual(row["platform"], "小红书")
            self.assertEqual(row["content_id"], "6a115b0e00000000360031a6")
            self.assertEqual(row["source_file"], "../feishu_exports/原生内容投稿.xlsx")


class DuplicatePolicyTests(unittest.TestCase):
    def test_duplicate_policy_keeps_different_accounts_separate(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "title": "同一个标题",
                    "account": "同花顺投资",
                    "manual_category": "资讯",
                    "spend": 100,
                    "impressions": 1000,
                    "activations": 10,
                    "first_pay_count": 2,
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "title": "同一个标题",
                    "account": "同花顺财富",
                    "manual_category": "资讯",
                    "spend": 200,
                    "impressions": 2000,
                    "activations": 20,
                    "first_pay_count": 4,
                },
            ]
        )

        analysis = analyze_canonical_frame(
            frame,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        )

        self.assertEqual(len(analysis.canonical), 2)
        self.assertEqual(set(analysis.canonical["account"]), {"同花顺投资", "同花顺财富"})

    def test_xiaohongshu_dedupes_by_content_id_sums_rows_and_keeps_tagged_title(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "小红书商业化",
                    "platform_group": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "note-dup",
                    "title": "5.7日段永平调仓买入泡泡玛特，头像也换了！",
                    "account": "投资号",
                    "manual_category": "资讯",
                    "spend": 100,
                    "impressions": 1000,
                    "clicks": 100,
                    "activations": 10,
                    "first_pay_count": 2,
                },
                {
                    "platform": "小红书商业化",
                    "platform_group": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "note-dup",
                    "title": "5.7日段永平调仓买入泡泡玛特，头像也换了！ #同花顺APP #同花顺资讯",
                    "account": "理财",
                    "manual_category": "资讯",
                    "spend": 120,
                    "impressions": 1200,
                    "clicks": 120,
                    "activations": 12,
                    "first_pay_count": 3,
                },
            ]
        )

        analysis = analyze_canonical_frame(
            frame,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        )

        self.assertEqual(len(analysis.canonical), 1)
        row = analysis.canonical.iloc[0]
        self.assertEqual(row["dedupe_key"], "小红书商业化::id::note-dup")
        self.assertEqual(row["merged_row_count"], 2)
        self.assertEqual(row["title"], "5.7日段永平调仓买入泡泡玛特，头像也换了！ #同花顺APP #同花顺资讯")
        self.assertEqual(row["spend"], 220)
        self.assertEqual(row["impressions"], 2200)
        self.assertEqual(row["clicks"], 220)
        self.assertEqual(row["activations"], 22)
        self.assertEqual(row["first_pay_count"], 5)
        self.assertFalse(bool(row["needs_manual_review"]))

    def test_bilibili_dedupes_by_bv_and_sums_repeated_identical_rows(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "B站",
                    "platform_group": "B站",
                    "channel": "B站",
                    "content_id": "BV1Same",
                    "title": "同一条B站内容",
                    "account": "同花顺投资",
                    "manual_category": "B站全部",
                    "spend": 50,
                    "impressions": 500,
                    "clicks": 50,
                    "activations": 5,
                    "first_pay_count": 1,
                },
                {
                    "platform": "B站",
                    "platform_group": "B站",
                    "channel": "B站",
                    "content_id": "BV1Same",
                    "title": "同一条B站内容",
                    "account": "同花顺投资",
                    "manual_category": "B站全部",
                    "spend": 50,
                    "impressions": 500,
                    "clicks": 50,
                    "activations": 5,
                    "first_pay_count": 1,
                },
            ]
        )

        analysis = analyze_canonical_frame(
            frame,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        )

        self.assertEqual(len(analysis.canonical), 1)
        row = analysis.canonical.iloc[0]
        self.assertEqual(row["dedupe_key"], "B站::id::BV1Same")
        self.assertEqual(row["merged_row_count"], 2)
        self.assertEqual(row["spend"], 100)
        self.assertEqual(row["impressions"], 1000)
        self.assertEqual(row["clicks"], 100)
        self.assertEqual(row["activations"], 10)
        self.assertEqual(row["first_pay_count"], 2)
        self.assertFalse(bool(row["needs_manual_review"]))

    def test_xiaohongshu_id_dedupe_keeps_channels_separate_and_flags_different_titles(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "小红书商业化",
                    "platform_group": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "note-shared",
                    "title": "第一版标题",
                    "account": "投资号",
                    "manual_category": "资讯",
                    "spend": 100,
                },
                {
                    "platform": "小红书商业化",
                    "platform_group": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "note-shared",
                    "title": "完全不同标题 #同花顺资讯",
                    "account": "投资号",
                    "manual_category": "资讯",
                    "spend": 120,
                },
                {
                    "platform": "小红书市场部",
                    "platform_group": "小红书",
                    "channel": "小红书市场部",
                    "content_id": "note-shared",
                    "title": "市场部同ID不跨渠道合并",
                    "account": "投资号",
                    "manual_category": "资讯",
                    "spend": 80,
                },
            ]
        )

        analysis = analyze_canonical_frame(
            frame,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        )

        self.assertEqual(len(analysis.canonical), 2)
        commercial = analysis.canonical[analysis.canonical["channel"].eq("小红书商业化")].iloc[0]
        market = analysis.canonical[analysis.canonical["channel"].eq("小红书市场部")].iloc[0]
        self.assertEqual(commercial["spend"], 220)
        self.assertEqual(commercial["title"], "完全不同标题 #同花顺资讯")
        self.assertTrue(bool(commercial["needs_manual_review"]))
        self.assertIn("同ID标题不一致", commercial["review_reasons"])
        self.assertEqual(market["spend"], 80)
        self.assertEqual(market["merged_row_count"], 1)

    def test_douyin_same_content_id_still_keeps_different_accounts_separate(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "aweme-shared",
                    "title": "同一个视频",
                    "account": "同花顺投资",
                    "manual_category": "资讯",
                    "spend": 100,
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "aweme-shared",
                    "title": "同一个视频",
                    "account": "同花顺财富",
                    "manual_category": "资讯",
                    "spend": 200,
                },
            ]
        )

        analysis = analyze_canonical_frame(
            frame,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        )

        self.assertEqual(len(analysis.canonical), 2)
        self.assertEqual(set(analysis.canonical["account"]), {"同花顺投资", "同花顺财富"})

    def test_duplicate_policy_sums_large_differences_and_reviews_close_values(self):
        large_difference = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "aweme-1",
                    "title": "同一条内容",
                    "account": "同花顺投资",
                    "manual_category": "资讯",
                    "spend": 100,
                    "impressions": 1000,
                    "activations": 10,
                    "first_pay_count": 2,
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "aweme-1",
                    "title": "同一条内容",
                    "account": "同花顺投资",
                    "manual_category": "资讯",
                    "spend": 180,
                    "impressions": 1800,
                    "activations": 18,
                    "first_pay_count": 4,
                },
            ]
        )
        close_values = large_difference.copy()
        close_values.loc[1, ["spend", "impressions", "activations", "first_pay_count"]] = [103, 1030, 10, 2]

        summed = analyze_canonical_frame(
            large_difference,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        ).canonical.iloc[0]
        reviewed = analyze_canonical_frame(
            close_values,
            "2026-05-19",
            "2026-05-25",
            category_matcher=lambda items, category_library, env_path: {},
        ).canonical.iloc[0]

        self.assertEqual(summed["spend"], 280)
        self.assertFalse(bool(summed["needs_manual_review"]))
        self.assertIn("spend=100 | 180->sum", summed["conflict_details"])
        self.assertEqual(reviewed["spend"], 100)
        self.assertTrue(bool(reviewed["needs_manual_review"]))
        self.assertIn("数值相近重复待审核", reviewed["review_reasons"])


class RecapSummaryTests(unittest.TestCase):
    def test_week_and_month_recap_fields_match_business_metrics(self):
        items = pd.DataFrame(
            [
                {
                    "channel": "抖音商业化",
                    "spend": 100,
                    "impressions": 1000,
                    "activations": 10,
                    "first_pay_count": 2,
                },
                {
                    "channel": "B站",
                    "spend": 300,
                    "impressions": 3000,
                    "activations": 30,
                    "first_pay_count": 3,
                },
            ]
        )

        weekly = build_recap_summary(items, period_level="week")
        monthly = build_recap_summary(items, period_level="month")

        self.assertEqual(
            list(weekly.columns),
            ["渠道", "消耗", "曝光量", "激活数", "激活成本", "付费", "付费成本"],
        )
        self.assertEqual(float(weekly[weekly["渠道"].eq("汇总")].iloc[0]["激活成本"]), 10)
        self.assertIn("大盘付费数据", monthly.columns)
        self.assertIn("大盘付费成本", monthly.columns)
        self.assertEqual(monthly[monthly["渠道"].eq("汇总")].iloc[0]["大盘付费数据"], "占位")
        self.assertEqual(float(monthly[monthly["渠道"].eq("汇总")].iloc[0]["原生内容曝光数"]), 4000)
        self.assertEqual(float(monthly[monthly["渠道"].eq("B站")].iloc[0]["消耗占比"]), 0.75)

    def test_report_exports_unified_recap_fields(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            canonical = pd.DataFrame(
                [
                    {
                        "channel": "抖音商业化",
                        "content_id": "dy-1",
                        "title": "测试内容",
                        "spend": 100,
                        "impressions": 1000,
                        "clicks": 10,
                        "activations": 10,
                        "first_pay_count": 2,
                        "content_category": "资讯",
                    }
                ]
            )

            _, analysis_xlsx, _, total_summary_xlsx = write_outputs(
                root,
                "2026-05-01",
                "2026-05-31",
                canonical,
                pd.DataFrame(),
                pd.DataFrame(),
                pd.DataFrame(),
                pd.DataFrame(),
                pd.DataFrame(),
                pd.DataFrame(),
                pd.DataFrame(),
            )

            analysis_workbook = load_workbook(analysis_xlsx, read_only=True)
            total_workbook = load_workbook(total_summary_xlsx, read_only=True)
            self.assertIn("复盘统一字段", analysis_workbook.sheetnames)
            self.assertIn("复盘统一字段", total_workbook.sheetnames)
            headers = [
                cell.value
                for cell in next(analysis_workbook["复盘统一字段"].iter_rows(min_row=1, max_row=1))
            ]
            self.assertIn("大盘付费数据", headers)
            self.assertIn("原生内容曝光数", headers)
