import os
from pathlib import Path
from contextlib import closing
import json
import sqlite3
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from ops_data_workflow.workflow import refresh_historical_source_periods, run_archived_workflow, run_workflow
from ops_data_workflow.pipeline import analyze_canonical_frame, analyze_input_dir
from ops_data_workflow.recap_dataset import build_cleaned_asset_table, build_content_recap_table, build_unanalyzable_summary
from ops_data_workflow.reference_tables import account_mapping_lookup, load_reference_tables, parse_period_from_raw_dir
from ops_data_workflow.cleaning_pipeline import split_channel_total_rows
from ops_data_workflow.top_asset_service import build_executable_top_content_pool
from ops_data_workflow.title_matching import normalized_title_key


CORE_ANALYSIS_SHEETS = ["清洗后素材表", "内容复盘表", "不可分析汇总", "匹配覆盖率", "已匹配账号类型分析", "未匹配归因"]


def _write_raw_fixture(raw_dir: Path) -> None:
    with pd.ExcelWriter(raw_dir / "B站.xlsx", engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {
                    "视频AVID": "av1",
                    "视频BVID": "bv1",
                    "视频标题": "实盘大赛冠军孙辉--370万到2000万的传奇交易之路",
                    "Up主mid": "1622777305",
                    "日期": "2026-04-01~2026-04-27",
                    "花费": 100.0,
                    "展示量": 10000,
                    "点击量": 500,
                    "应用激活数": 20,
                    "应用内付费": 4,
                }
            ]
        ).to_excel(writer, sheet_name="sheet1", index=False)

    with pd.ExcelWriter(raw_dir / "小红书商业化.xlsx", engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {
                    "时间": "2026-04-01~2026-04-27",
                    "标题": "存储芯片板块再度爆发",
                    "笔记ID": "note-1",
                    "发布作者": "同花顺理财",
                    "类型": "图文",
                    "内容分类": "热点行情",
                    "消费": 60.0,
                    "展现量": 6000,
                    "点击量": 300,
                    "激活数": 12,
                    "首次付费次数": 2,
                    "内容类型": "",
                },
                {
                    "时间": "2026-04-01~2026-04-27",
                    "标题": "给短线交易者的完美范例 #股友说",
                    "笔记ID": "note-2",
                    "发布作者": "同顺股民社区",
                    "类型": "视频",
                    "内容分类": "",
                    "消费": 40.0,
                    "展现量": 4000,
                    "点击量": 100,
                    "激活数": 4,
                    "首次付费次数": 1,
                    "内容类型": "",
                },
            ]
        ).to_excel(writer, sheet_name="kos账户投放数据", index=False)
        pd.DataFrame(
            [
                {"笔记ID": "note-1", "账号": "同花顺投资", "内容类型": "资讯"},
            ]
        ).to_excel(writer, sheet_name="内容表格", index=False, startrow=1)

    with pd.ExcelWriter(raw_dir / "抖音商业化.xlsx", engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {
                    "视频标题": "什么样的人能成为交易高手？ #股友说 #股民",
                    "视频id": "7310000000000000001",
                    "素材ID": "mat-1",
                    "账号": "同花顺投资",
                    "消耗": 200.0,
                    "展示数": 20000,
                    "点击数": 800,
                    "激活数": 50,
                    "付费次数": 20,
                    "内容类型": "股友说",
                },
                {
                    "视频标题": "股市是仅次于高考最公平的竞争",
                    "视频id": "7310000000000000002",
                    "素材ID": "mat-2",
                    "账号": "同花顺投资",
                    "消耗": 100.0,
                    "展示数": 10000,
                    "点击数": 400,
                    "激活数": 20,
                    "付费次数": 8,
                    "内容类型": "资讯",
                },
            ]
        ).to_excel(writer, sheet_name="Sheet2", index=False)

    with pd.ExcelWriter(raw_dir / "抖音市场部.xlsx", engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {
                    "创建时间": "2026-04-01",
                    "素材ID": "mat-3",
                    "视频标题": "是天才就来同花顺证明给我看 #同花顺进行曲",
                    "视频id": "7310000000000000003",
                    "账号": "同花顺投资",
                    "消耗": 120.0,
                    "展示数": 12000,
                    "激活数": 24,
                    "付费次数": 10,
                    "内容类型": "",
                }
            ]
        ).to_excel(writer, sheet_name="Sheet2", index=False)

    total = pd.DataFrame(
        [
            {"渠道": "B站", "消耗": "50", "激活": "10", "付费": "2"},
            {"渠道": "小红书商业化", "消耗": "100", "激活": "16", "付费": "3"},
            {"渠道": "抖音商业化", "消耗": "150", "激活": "35", "付费": "14"},
            {"渠道": "抖音市场部", "消耗": "60", "激活": "12", "付费": "5"},
        ]
    )
    with pd.ExcelWriter(raw_dir / "四月消耗，占比等总数据.xlsx", engine="openpyxl") as writer:
        total.to_excel(writer, sheet_name="Sheet1", index=False, startrow=5, startcol=8)


def _feishu_ledger_fixture() -> pd.DataFrame:
    frame = pd.DataFrame(
        [
            {
                "platform": "抖音",
                "published_date": "2026-04-01",
                "content_url": "https://www.douyin.com/video/7310000000000000001",
                "content_id": "7310000000000000001",
                "account": "同花顺投资",
                "title": "什么样的人能成为交易高手？ #股友说 #股民",
                "tags": "#股友说 #股民",
                "raw_content_type": "股友说",
                "category_l1": "",
                "category_l2": "股友说",
                "bilibili_content_type": "",
                "content_type": "股友说",
                "content_type_review": "",
                "filter_status": "",
                "source_file": "harvester_feishu",
                "source_sheet": "抖音渠道",
                "source_row": 2,
                "title_key": normalized_title_key("什么样的人能成为交易高手？ #股友说 #股民"),
                "title_key_no_tags": normalized_title_key("什么样的人能成为交易高手？"),
            },
            {
                "platform": "抖音",
                "published_date": "2026-04-02",
                "content_url": "https://www.douyin.com/video/7310000000000000002",
                "content_id": "7310000000000000002",
                "account": "同花顺投资",
                "title": "股市是仅次于高考最公平的竞争",
                "tags": "#资讯",
                "raw_content_type": "资讯",
                "category_l1": "",
                "category_l2": "资讯",
                "bilibili_content_type": "",
                "content_type": "资讯",
                "content_type_review": "",
                "filter_status": "",
                "source_file": "harvester_feishu",
                "source_sheet": "抖音渠道",
                "source_row": 3,
                "title_key": normalized_title_key("股市是仅次于高考最公平的竞争"),
                "title_key_no_tags": normalized_title_key("股市是仅次于高考最公平的竞争"),
            },
            {
                "platform": "抖音",
                "published_date": "2026-04-03",
                "content_url": "https://www.douyin.com/video/7310000000000000003",
                "content_id": "7310000000000000003",
                "account": "同花顺投资",
                "title": "是天才就来同花顺证明给我看 #同花顺进行曲",
                "tags": "#同花顺进行曲",
                "raw_content_type": "品牌活动",
                "category_l1": "",
                "category_l2": "品牌活动",
                "bilibili_content_type": "",
                "content_type": "品牌活动",
                "content_type_review": "",
                "filter_status": "",
                "source_file": "harvester_feishu",
                "source_sheet": "抖音渠道",
                "source_row": 4,
                "title_key": normalized_title_key("是天才就来同花顺证明给我看 #同花顺进行曲"),
                "title_key_no_tags": normalized_title_key("是天才就来同花顺证明给我看"),
            },
        ]
    )
    frame.attrs["feishu_snapshot"] = {
        "enabled": True,
        "total_rows": int(len(frame)),
        "platform_counts": {"抖音": int(len(frame))},
        "sheet_row_counts": {"dySheet": int(len(frame))},
        "field_completeness": {"content_id": 1.0},
        "warnings": [],
    }
    frame.attrs["source_files"] = set()
    return frame


def _empty_feishu_ledger_fixture() -> pd.DataFrame:
    frame = pd.DataFrame(
        columns=[
            "platform",
            "published_date",
            "content_url",
            "content_id",
            "account",
            "title",
            "tags",
            "raw_content_type",
            "category_l1",
            "category_l2",
            "bilibili_content_type",
            "content_type",
            "content_type_review",
            "filter_status",
            "source_file",
            "source_sheet",
            "source_row",
            "title_key",
            "title_key_no_tags",
        ]
    )
    frame.attrs["feishu_enabled"] = False
    frame.attrs["ledger_warnings"] = ["缺少飞书配置"]
    frame.attrs["source_files"] = set()
    frame.attrs["feishu_snapshot"] = {
        "enabled": False,
        "total_rows": 0,
        "platform_counts": {},
        "sheet_row_counts": {},
        "field_completeness": {},
        "warnings": ["缺少飞书配置"],
    }
    return frame


def _remove_total_fixture(raw_dir: Path) -> None:
    (raw_dir / "四月消耗，占比等总数据.xlsx").unlink()


def _looks_like_english_field_name(value: object) -> bool:
    text = "" if value is None else str(value)
    return bool("__" in text or "_" in text or text in {"channel", "source_file", "canonical_column"})


class WorkflowTests(unittest.TestCase):
    def test_workflow_reads_csv_sources_and_builds_core_unanalyzable_tables(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            pd.DataFrame(
                [
                        {
                            "视频AVID": "av-csv",
                            "视频BVID": "bv-csv",
                            "视频标题": "B站财经内容",
                            "Up主mid": "1622777305",
                            "花费": 80,
                        "展示量": 8000,
                        "点击量": 400,
                        "应用激活数": 16,
                        "应用内付费": 4,
                        "素材中心id": "mat-b",
                    }
                ]
            ).to_csv(raw_dir / "B站.csv", index=False, encoding="utf-8-sig")
            pd.DataFrame(
                [
                        {
                            "视频标题": "抖音投流内容",
                            "视频id": "dy-csv",
                            "素材ID": "mat-dy",
                            "账号": "同花顺投资",
                            "消耗": 120,
                        "展示数": 10000,
                        "点击数": 500,
                        "激活数": 30,
                        "付费次数": 6,
                        "内容类型": "热点行情",
                    }
                ]
            ).to_csv(raw_dir / "抖音商业化.csv", index=False, encoding="utf-8-sig")

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            self.assertEqual(len(result.canonical), 2)
            platform_summary = result.platform_summary.set_index("channel")
            self.assertEqual(set(platform_summary.index), {"B站市场部", "抖音商业化"})
            self.assertAlmostEqual(platform_summary.loc["B站市场部", "spend"], 80.0)
            self.assertAlmostEqual(platform_summary.loc["抖音商业化", "spend"], 120.0)
            channel_summary = result.channel_summary.set_index("channel")
            self.assertEqual(set(channel_summary.index), {"B站市场部", "抖音商业化"})
            self.assertAlmostEqual(channel_summary.loc["B站市场部", "activations"], 16.0)
            self.assertAlmostEqual(channel_summary.loc["抖音商业化", "first_pay_count"], 6.0)
            self.assertTrue(result.platform_category_summary.empty)
            self.assertEqual(len(result.cleaned_asset_table), 2)
            self.assertTrue(result.content_recap_table.empty)
            summary = result.unanalyzable_summary.set_index("渠道")
            self.assertAlmostEqual(summary.loc["B站市场部", "不可分析消耗"], 80.0)
            self.assertAlmostEqual(summary.loc["抖音商业化", "不可分析消耗"], 120.0)

    def test_bilibili_aggregate_export_without_title_is_ingested_with_xiaohongshu(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "B站视频投放数据-5.15.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "视频bvid": "BV14Vo5BFE1w",
                            "Up主mid": "1622777305",
                            "求和项:总花费": 589.62,
                            "求和项:展示量": 33224,
                            "求和项:点击量": 187,
                            "求和项:应用激活数": 14,
                            "求和项:应用内首次付费次数": 3,
                        },
                        {
                            "视频bvid": "BV17EoKB9E7e",
                            "Up主mid": "1622777305",
                            "求和项:总花费": 147.84,
                            "求和项:展示量": 9547,
                            "求和项:点击量": 299,
                            "求和项:应用激活数": 3,
                            "求和项:应用内首次付费次数": 1,
                        },
                    ]
                ).to_excel(writer, sheet_name="Sheet1", index=False)
            with pd.ExcelWriter(raw_dir / "小红书账号投放数据（2026.05.08-05.14）.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "时间": "2026-05-08~2026-05-14",
                            "标题": "小红书内容",
                            "笔记ID": "note-1",
                            "发布作者": "同花顺理财",
                            "类型": "图文",
                            "内容分类": "热点行情",
                            "消费": 60.0,
                            "展现量": 6000,
                            "点击量": 300,
                            "激活数": 12,
                            "首次付费次数": 2,
                        }
                    ]
                ).to_excel(writer, sheet_name="02户", index=False)

            empty_env = Path(tmp) / "empty.env"
            empty_env.write_text("", encoding="utf-8")
            with patch.dict(os.environ, {key: "" for key in os.environ if key.startswith("FEISHU_")}, clear=False):
                result = analyze_input_dir(
                    raw_dir,
                    "2026-05-08",
                    "2026-05-14",
                    category_matcher=lambda items, category_library, env_path: {},
                    env_path=empty_env,
                )

            self.assertEqual(result.canonical["channel"].value_counts().to_dict(), {"B站市场部": 2, "小红书商业化": 1})
            bilibili = result.canonical[result.canonical["channel"].eq("B站市场部")].sort_values("content_id")
            first = bilibili.iloc[0]
            self.assertEqual(first["content_id"], "BV14Vo5BFE1w")
            self.assertIsInstance(first["title"], str)
            self.assertAlmostEqual(first["spend"], 589.62)
            self.assertAlmostEqual(first["activations"], 14.0)
            self.assertAlmostEqual(first["first_pay_count"], 3.0)
            self.assertEqual(first["content_form"], "视频")
            self.assertEqual(first["content_category"], "")
            self.assertEqual(first["category_l2"], "")
            self.assertIsInstance(first["category_l3"], str)

    def test_bilibili_impression_aliases_are_ingested_by_keyword(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "B站.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "视频AVID": "av-alias",
                            "视频BVID": "BVImpressionAlias",
                            "视频标题": "B站展示量别名内容",
                            "Up主mid": "1622777305",
                            "花费": 100.0,
                            "视频展示量": 4321,
                            "曝光转化率": 0.12,
                            "千次展示费用": 23.5,
                            "点击量": 321,
                            "应用激活数": 12,
                            "激活成本": 8.3,
                            "应用内首次付费次数": 4,
                            "付费成本": 25.0,
                        }
                    ]
                ).to_excel(writer, sheet_name="Sheet1", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = result.canonical.iloc[0]
            self.assertEqual(row["platform"], "B站")
            self.assertEqual(row["channel"], "B站市场部")
            self.assertAlmostEqual(row["impressions"], 4321.0)
            self.assertAlmostEqual(row["clicks"], 321.0)
            self.assertAlmostEqual(row["activations"], 12.0)
            self.assertAlmostEqual(row["first_pay_count"], 4.0)

    def test_xiaohongshu_commercial_metric_count_aliases_are_ingested(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "小红书商业化.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "小红书投放内容",
                            "笔记ID": "note-app-activation",
                            "发布作者": "同花顺理财",
                            "内容分类": "产品科普",
                            "消费": 100.0,
                            "展现量": 1000,
                            "点击量": 100,
                            "APP激活数": 9,
                            "首次付费数": 3,
                        }
                    ]
                ).to_excel(writer, sheet_name="kos账号笔记投放数据", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = result.canonical.iloc[0]
            self.assertEqual(row["channel"], "小红书商业化")
            self.assertAlmostEqual(row["activations"], 9.0)
            self.assertAlmostEqual(row["first_pay_count"], 3.0)
            channel_summary = result.channel_summary.set_index("channel")
            self.assertIn("小红书商业化", channel_summary.index)
            self.assertAlmostEqual(channel_summary.loc["小红书商业化", "activations"], 9.0)
            self.assertAlmostEqual(channel_summary.loc["小红书商业化", "first_pay_count"], 3.0)
            summary = build_unanalyzable_summary(build_cleaned_asset_table(result.canonical)).set_index("渠道")
            self.assertEqual(int(summary.loc["小红书商业化", "总素材数"]), 1)
            self.assertAlmostEqual(summary.loc["小红书商业化", "不可分析消耗"], 100.0)

    def test_xiaohongshu_commercial_seven_day_pay_count_alias_is_ingested(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "小红书商业化.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "小红书七日付费内容",
                            "笔记ID": "note-seven-day-pay",
                            "发布作者": "同花顺理财",
                            "内容分类": "产品科普",
                            "消费": 100.0,
                            "展现量": 1000,
                            "点击量": 100,
                            "APP激活数": 9,
                            "7日付费次数": 5,
                        }
                    ]
                ).to_excel(writer, sheet_name="kos账号笔记投放数据", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = result.canonical.iloc[0]
            self.assertEqual(row["channel"], "小红书商业化")
            self.assertAlmostEqual(row["first_pay_count"], 5.0)
            channel_summary = result.channel_summary.set_index("channel")
            self.assertIn("小红书商业化", channel_summary.index)
            self.assertAlmostEqual(channel_summary.loc["小红书商业化", "first_pay_count"], 5.0)
            summary = build_unanalyzable_summary(build_cleaned_asset_table(result.canonical)).set_index("渠道")
            self.assertEqual(int(summary.loc["小红书商业化", "总素材数"]), 1)
            self.assertAlmostEqual(summary.loc["小红书商业化", "不可分析消耗"], 100.0)

    def test_xiaohongshu_commercial_metric_rates_are_not_ingested_as_counts(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "小红书商业化.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "小红书投放内容",
                            "笔记ID": "note-rate-only",
                            "发布作者": "同花顺理财",
                            "内容分类": "产品科普",
                            "消费": 100.0,
                            "展现量": 1000,
                            "点击量": 100,
                            "激活成本": 12.3,
                            "激活率": 0.12,
                            "首次付费成本": 33.3,
                            "首次付费率": 0.25,
                        }
                    ]
                ).to_excel(writer, sheet_name="kos账号笔记投放数据", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = result.canonical.iloc[0]
            self.assertTrue(pd.isna(row["activations"]))
            self.assertTrue(pd.isna(row["first_pay_count"]))

    def test_xiaohongshu_account_filter_config_no_longer_excludes_raw_excel_input(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            config_path = tmp_path / "config" / "account_filters.yml"
            config_path.parent.mkdir()
            config_path.write_text(
                """
xiaohongshu:
  include_accounts:
    - 股民社区
    - 研习社
  aliases:
    同顺股民社区: 股民社区
    同花顺研习社: 研习社
  exclude_blank: true
""".strip(),
                encoding="utf-8",
            )
            with pd.ExcelWriter(raw_dir / "小红书商业化.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "股民社区内容",
                            "笔记ID": "note-included-1",
                            "发布作者": "同顺股民社区",
                            "内容分类": "投教",
                            "消费": 10.0,
                            "激活数": 1,
                            "首次付费次数": 1,
                        },
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "研习社内容",
                            "笔记ID": "note-included-2",
                            "发布作者": "同花顺研习社",
                            "内容分类": "投教",
                            "消费": 20.0,
                            "激活数": 2,
                            "首次付费次数": 1,
                        },
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "APP内容",
                            "笔记ID": "note-excluded-app",
                            "发布作者": "同花顺APP",
                            "内容分类": "产品",
                            "消费": 999.0,
                            "激活数": 99,
                            "首次付费次数": 9,
                        },
                        {
                            "时间": "2026-05-15~2026-05-21",
                            "标题": "空账号内容",
                            "笔记ID": "note-included-blank",
                            "发布作者": "",
                            "内容分类": "产品",
                            "消费": 888.0,
                            "激活数": 88,
                            "首次付费次数": 8,
                        },
                    ]
                ).to_excel(writer, sheet_name="kos账号笔记投放数据", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-15",
                "2026-05-21",
                account_filters_path=config_path,
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(
                set(result.canonical["content_id"]),
                {"note-included-1", "note-included-2", "note-excluded-app", "note-included-blank"},
            )
            self.assertEqual(set(result.canonical["account"].fillna("")), {"", "同顺股民社区", "同花顺APP", "同花顺研习社"})
            self.assertAlmostEqual(result.canonical["spend"].sum(), 1917.0)
            self.assertEqual(set(result.channel_summary["channel"]), {"小红书商业化"})
            summary = build_unanalyzable_summary(build_cleaned_asset_table(result.canonical)).set_index("渠道")
            self.assertAlmostEqual(summary.loc["小红书商业化", "不可分析消耗"], 1917.0)
            self.assertTrue(result.account_filter_details.empty)
            self.assertNotIn("小红书账号过滤排除行数", set(result.preprocessing_report["metric"]))

    def test_xiaohongshu_account_filter_config_no_longer_excludes_cleaned_replay_frame(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "account_filters.yml"
            config_path.write_text(
                """
xiaohongshu:
  include_accounts:
    - 股民社区
    - 研习社
  aliases:
    同顺股民社区: 股民社区
    同花顺研习社: 研习社
  exclude_blank: true
""".strip(),
                encoding="utf-8",
            )
            canonical = pd.DataFrame(
                [
                    {
                        "channel": "小红书商业化",
                        "content_id": "note-included-1",
                        "title": "股民社区内容",
                        "account": "同顺股民社区",
                        "manual_category": "投教",
                        "spend": 10.0,
                        "activations": 1,
                        "first_pay_count": 1,
                    },
                    {
                        "channel": "小红书商业化",
                        "content_id": "note-included-2",
                        "title": "研习社内容",
                        "account": "同花顺研习社",
                        "manual_category": "投教",
                        "spend": 20.0,
                        "activations": 2,
                        "first_pay_count": 1,
                    },
                    {
                        "channel": "小红书商业化",
                        "content_id": "note-excluded-etf",
                        "title": "ETF内容",
                        "account": "同花顺ETF",
                        "manual_category": "产品",
                        "spend": 999.0,
                        "activations": 99,
                        "first_pay_count": 9,
                    },
                    {
                        "channel": "小红书商业化",
                        "content_id": "note-included-blank",
                        "title": "空账号内容",
                        "account": "",
                        "manual_category": "产品",
                        "spend": 888.0,
                        "activations": 88,
                        "first_pay_count": 8,
                    },
                    {
                        "channel": "抖音商业化",
                        "content_id": "dy-kept",
                        "title": "抖音内容",
                        "account": "",
                        "manual_category": "投教",
                        "spend": 50.0,
                        "activations": 5,
                        "first_pay_count": 1,
                    },
                ]
            )

            result = analyze_canonical_frame(
                canonical,
                "2026-05-15",
                "2026-05-21",
                account_filters_path=config_path,
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(
                set(result.canonical["content_id"]),
                {"note-included-1", "note-included-2", "note-excluded-etf", "note-included-blank", ""},
            )
            douyin = result.canonical[result.canonical["channel"].eq("抖音商业化")].iloc[0]
            self.assertEqual(douyin["material_id"], "dy-kept")
            self.assertEqual(douyin["ad_material_id"], "dy-kept")
            self.assertEqual(
                set(result.canonical[result.canonical["channel"].eq("小红书商业化")]["account"].fillna("")),
                {"", "同顺股民社区", "同花顺ETF", "同花顺研习社"},
            )
            self.assertTrue(result.account_filter_details.empty)
            self.assertEqual(set(result.channel_summary["channel"]), {"小红书商业化", "抖音商业化"})
            summary = build_unanalyzable_summary(build_cleaned_asset_table(result.canonical)).set_index("渠道")
            self.assertAlmostEqual(summary.loc["小红书商业化", "不可分析消耗"], 1917.0)

    def test_xiaohongshu_market_reads_all_metric_sheets(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "小红书市场部.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "笔记/素材ID": "note-new",
                            "标题": "新户内容",
                            "账号": "问财",
                            "消费": 10.0,
                            "展现量": 100,
                            "点击量": 10,
                            "激活数": 2,
                            "首次付费次数": 1,
                        }
                    ]
                ).to_excel(writer, sheet_name="新户", index=False)
                pd.DataFrame(
                    [
                        {
                            "笔记/素材ID": "note-old-1",
                            "标题": "老户内容1",
                            "账号": "投资号",
                            "消费": 20.0,
                            "展现量": 200,
                            "点击量": 20,
                            "激活数": 4,
                            "首次付费次数": 2,
                        },
                        {
                            "笔记/素材ID": "note-old-2",
                            "标题": "老户内容2",
                            "账号": "财经号",
                            "消费": 30.0,
                            "展现量": 300,
                            "点击量": 30,
                            "激活数": 6,
                            "首次付费次数": 3,
                        },
                    ]
                ).to_excel(writer, sheet_name="老户", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-01",
                "2026-05-31",
                category_matcher=lambda items, category_library, env_path: {},
            )

            xhs = result.canonical[result.canonical["channel"].eq("小红书市场部")]
            self.assertEqual(len(xhs), 3)
            self.assertEqual(set(xhs["source_sheet"]), {"新户", "老户"})
            self.assertAlmostEqual(xhs["spend"].sum(), 60.0)
            self.assertAlmostEqual(xhs["activations"].sum(), 12.0)
            self.assertAlmostEqual(xhs["first_pay_count"].sum(), 6.0)

    def test_xiaohongshu_market_link_only_rows_keep_link_out_of_title(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            content_url = "https://www.xiaohongshu.com/explore/note-link-only?xsec_source=pc_ad_export"
            with pd.ExcelWriter(raw_dir / "小红书市场部.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "笔记/素材ID": "note-link-only",
                            "笔记/素材链接": content_url,
                            "消费": 10.0,
                            "展现量": 100,
                            "点击量": 10,
                            "激活数": 2,
                            "首次付费次数": 1,
                        }
                    ]
                ).to_excel(writer, sheet_name="新户", index=False)

            ledger = pd.DataFrame()
            ledger.attrs["feishu_enabled"] = False
            ledger.attrs["ledger_warnings"] = []
            ledger.attrs["source_files"] = set()
            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=ledger):
                result = analyze_input_dir(
                    raw_dir,
                    "2026-05-01",
                    "2026-05-31",
                    category_matcher=lambda items, category_library, env_path: {},
                )

            row = result.canonical[result.canonical["content_id"].eq("note-link-only")].iloc[0]
            self.assertEqual(row["content_url"], content_url)
            self.assertEqual(row["title"], "")

    def test_unmatched_assets_do_not_call_category_matcher(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            pd.DataFrame(
                [
                    {
                        "视频标题": "一个暂时无法从标题判断的内容",
                        "视频id": "dy-unknown",
                        "素材ID": "mat-unknown",
                        "账号": "同花顺投资",
                        "消耗": 100.0,
                        "展示数": 1000,
                        "点击数": 100,
                        "激活数": 10,
                        "付费次数": 2,
                        "内容类型": "",
                    }
                ]
            ).to_csv(raw_dir / "抖音市场部.csv", index=False, encoding="utf-8-sig")

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_empty_feishu_ledger_fixture()):
                result = analyze_input_dir(
                    raw_dir,
                    "2026-04-01",
                    "2026-04-27",
                    category_matcher=lambda items, category_library, env_path: (_ for _ in ()).throw(
                        AssertionError("不可分析素材不应进入 AI 分类")
                    ),
                    env_path=Path(tmp) / ".env",
                )

            inferred = result.canonical[result.canonical["material_id"].eq("mat-unknown")].iloc[0]
            self.assertEqual(inferred["content_id"], "")
            self.assertEqual(inferred["work_id"], "")
            self.assertEqual(inferred["manual_category"], "")
            self.assertEqual(inferred["ai_category"], "")
            self.assertEqual(inferred["content_category"], "")
            self.assertEqual(inferred["category_status"], "不可分析")
            self.assertEqual(inferred["analysis_status"], "不可分析")
            self.assertEqual(inferred["unanalyzable_reason"], "飞书台账缺失候选")
            asset_table = build_cleaned_asset_table(result.canonical)
            self.assertTrue(build_content_recap_table(asset_table).empty)

    def test_title_tags_and_high_spend_rules_do_not_analyze_unmatched_assets(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "dy-trading-mindset",
                    "title": "股市是仅次于高考，最公平的竞争 #同花顺 #交易心法",
                    "spend": 262031.87,
                    "manual_category": "",
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "dy-unknown-low",
                    "title": "全国在校大学生博主招募，有意私信",
                    "spend": 20.0,
                    "manual_category": "",
                },
            ]
        )

        result = analyze_canonical_frame(
            frame,
            "2026-05-01",
            "2026-05-31",
            category_matcher=lambda items, category_library, env_path: (_ for _ in ()).throw(
                AssertionError("不可分析素材不应进入 AI 分类")
            ),
        )

        canonical = result.canonical.set_index("material_id")
        self.assertEqual(canonical.loc["dy-trading-mindset", "content_id"], "")
        self.assertEqual(canonical.loc["dy-trading-mindset", "work_id"], "")
        self.assertEqual(canonical.loc["dy-trading-mindset", "content_category"], "")
        self.assertEqual(canonical.loc["dy-trading-mindset", "category_status"], "不可分析")
        self.assertEqual(canonical.loc["dy-unknown-low", "content_id"], "")
        self.assertEqual(canonical.loc["dy-unknown-low", "content_category"], "")
        self.assertEqual(canonical.loc["dy-unknown-low", "category_status"], "不可分析")
        top = result.top_content_items.set_index("material_id")
        self.assertIn("dy-trading-mindset", top.index)
        self.assertEqual(top.loc["dy-trading-mindset", "content_category"], "")
        self.assertEqual(canonical.loc["dy-trading-mindset", "analysis_status"], "不可分析")
        self.assertIn("dy-unknown-low", top.index)
        self.assertEqual(top.loc["dy-unknown-low", "content_category"], "")
        self.assertTrue(build_executable_top_content_pool(result.canonical).empty)

    def test_workflow_standardizes_sources_and_preserves_pending_categories(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_feishu_ledger_fixture()):
                result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            canonical = result.canonical
            self.assertEqual(len(canonical), 6)
            self.assertIn("source_file", canonical.columns)
            self.assertIn("work_id", canonical.columns)
            self.assertIn("work_url", canonical.columns)
            self.assertIn("analysis_status", canonical.columns)
            self.assertEqual(set(canonical["primary_category"].fillna("").astype(str)), {""})
            self.assertEqual(set(canonical["category_l1"].fillna("").astype(str)), {""})
            bilibili = canonical[canonical["content_id"].eq("bv1")].iloc[0]
            self.assertEqual(bilibili["content_form"], "视频")
            self.assertEqual(bilibili["content_category"], "")
            self.assertEqual(bilibili["category_l2"], "")
            self.assertEqual(bilibili["analysis_status"], "不可分析")
            self.assertEqual(bilibili["unanalyzable_reason"], "缺少作品ID或链接")
            self.assertEqual(
                canonical.loc[canonical["content_id"].eq("note-1"), "manual_category"].iloc[0],
                "热点行情",
            )
            self.assertEqual(
                canonical.loc[canonical["content_id"].eq("note-1"), "content_category"].iloc[0],
                "",
            )
            self.assertEqual(
                canonical.loc[canonical["content_id"].eq("note-2"), "category_status"].iloc[0],
                "不可分析",
            )
            self.assertEqual(len(result.pending_categories), 0)
            self.assertEqual(len(result.content_recap_table), 3)
            self.assertEqual(len(result.cleaned_asset_table), 6)

    def test_single_row_channel_total_is_excluded_from_top_content_items(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "row:d926474c43af",
                    "material_id": "row:d926474c43af",
                    "title": "小红书（商业化） 第3行",
                    "source_file": "小红书（商业化）.xlsx",
                    "source_sheet": "Sheet1",
                    "source_row": 3,
                    "spend": 104792,
                    "impressions": 2020000,
                    "activations": 2077,
                    "first_pay_count": 541,
                }
            ]
        )

        result = analyze_canonical_frame(frame, "2026-05-26", "2026-06-04")
        detail, totals = split_channel_total_rows(result.canonical)

        self.assertTrue(detail.empty)
        self.assertEqual(len(totals), 1)
        self.assertEqual(set(result.channel_summary["channel"]), {"小红书商业化"})
        self.assertTrue(result.top_content_items.empty)

    def test_category_completion_keeps_uploaded_l1_when_match_l1_is_blank(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "65f00000abcdef",
                    "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                    "title": "源数据标题",
                    "category_l1": "投教",
                    "manual_category": "理财方法",
                    "content_category": "理财方法",
                    "spend": 100,
                    "impressions": 1000,
                }
            ]
        )
        ledger = pd.DataFrame(
            [
                {
                    "platform": "小红书",
                    "content_id": "65f00000abcdef",
                    "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                    "title": "飞书标题",
                    "category_l1": "",
                    "category_l2": "理财方法",
                    "content_type": "理财方法",
                }
            ]
        )

        result = analyze_canonical_frame(frame, "2026-05-01", "2026-05-07", content_ledger=ledger)
        row = result.canonical.iloc[0]

        self.assertEqual(row["match_status"], "已匹配")
        self.assertEqual(row["category_l1"], "投教")
        self.assertEqual(row["category_l2"], "理财方法")

    def test_workflow_uses_independent_channels_without_l1_categories(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_feishu_ledger_fixture()):
                result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)
            canonical = result.canonical

            self.assertEqual(
                set(canonical["platform"]),
                {"B站", "小红书", "抖音"},
            )
            self.assertEqual(
                set(canonical["channel"]),
                {"B站市场部", "小红书商业化", "抖音商业化", "抖音市场部"},
            )
            self.assertEqual(
                set(canonical.loc[canonical["platform"].str.contains("抖音"), "platform_group"]),
                {"抖音"},
            )
            self.assertEqual(set(canonical["category_l1"].fillna("").astype(str)), {""})
            self.assertEqual(set(canonical["primary_category"].fillna("").astype(str)), {""})
            self.assertTrue(canonical["category_l2"].equals(canonical["content_category"]))
            self.assertTrue(canonical["category_source"].equals(canonical["category_status"]))
            self.assertEqual(set(canonical["analysis_status"]), {"不可分析", "可分析"})
            self.assertEqual(
                set(canonical.loc[canonical["analysis_status"].eq("可分析"), "channel"]),
                {"抖音商业化", "抖音市场部"},
            )
            self.assertEqual(
                set(result.channel_summary["channel"]),
                {"B站市场部", "小红书商业化", "抖音商业化", "抖音市场部"},
            )
            self.assertEqual(len(result.top_content_items), 6)
            self.assertEqual(
                set(build_executable_top_content_pool(result.canonical)["channel"]),
                {"抖音商业化", "抖音市场部"},
            )

    def test_workflow_resolves_known_bilibili_mid_to_account_name(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_feishu_ledger_fixture()):
                result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)
            canonical = result.canonical

            bilibili = canonical[canonical["content_id"].eq("bv1")].iloc[0]
            xhs = canonical[canonical["content_id"].eq("note-1")].iloc[0]
            douyin_market = canonical[canonical["material_id"].eq("mat-3")].iloc[0]

            self.assertEqual(bilibili["account_id"], "1622777305")
            self.assertEqual(bilibili["account"], "同花顺投资")
            self.assertEqual(bilibili["author"], "同花顺投资")
            self.assertEqual(xhs["account"], "同花顺理财")
            self.assertEqual(xhs["author"], "同花顺理财")
            self.assertEqual(douyin_market["account"], "同花顺投资")

    def test_reference_tables_initialize_account_mapping_and_period_from_raw_dir(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_period_dir = tmp_path / "data" / "raw" / "20260401-20260427"
            raw_period_dir.mkdir(parents=True)

            self.assertEqual(parse_period_from_raw_dir(raw_period_dir), ("2026-04-01", "2026-04-27"))

            references = load_reference_tables(tmp_path / "config" / "reference_tables.xlsx")

            self.assertTrue((tmp_path / "config" / "reference_tables.xlsx").exists())
            account_mapping = references.account_mapping.set_index(["channel", "source_account_id"])
            self.assertEqual(account_mapping.loc[("B站市场部", "1622777305"), "account"], "同花顺投资")
            self.assertEqual(account_mapping.loc[("B站商业化", "1622777305"), "account"], "同花顺投资")
            self.assertIn("账号映射表", references.tables)
            self.assertIn("字段映射表", references.tables)

    def test_reference_tables_workbook_uses_chinese_headers_and_is_still_readable(self):
        with TemporaryDirectory() as tmp:
            reference_path = Path(tmp) / "config" / "reference_tables.xlsx"

            references = load_reference_tables(reference_path)
            workbook = load_workbook(reference_path, read_only=True)

            account_headers = [cell.value for cell in next(workbook["账号映射表"].iter_rows(max_row=1))]
            field_headers = [cell.value for cell in next(workbook["字段映射表"].iter_rows(max_row=1))]
            workbook.close()

            self.assertEqual(account_headers, ["渠道", "来源账号ID", "来源账号名", "实际账号", "映射来源", "说明"])
            self.assertIn("标准字段", field_headers)
            self.assertNotIn("source_account_id", account_headers)
            self.assertNotIn("canonical_column", field_headers)
            lookup = account_mapping_lookup(references.account_mapping)
            self.assertEqual(lookup[("B站市场部", "1622777305")]["account"], "同花顺投资")
            self.assertEqual(lookup[("B站商业化", "1622777305")]["account"], "同花顺投资")

    def test_reference_tables_migrate_legacy_bilibili_account_mapping_rows(self):
        with TemporaryDirectory() as tmp:
            reference_path = Path(tmp) / "config" / "reference_tables.xlsx"
            reference_path.parent.mkdir(parents=True)
            with pd.ExcelWriter(reference_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "渠道": "B站",
                            "来源账号ID": "1622777305",
                            "来源账号名": "",
                            "实际账号": "同花顺投资",
                            "映射来源": "历史维护",
                            "说明": "旧表",
                        }
                    ]
                ).to_excel(writer, sheet_name="账号映射表", index=False)

            references = load_reference_tables(reference_path)

            account_mapping = references.account_mapping.set_index(["channel", "source_account_id"])
            self.assertEqual(account_mapping.loc[("B站市场部", "1622777305"), "account"], "同花顺投资")
            self.assertEqual(account_mapping.loc[("B站商业化", "1622777305"), "account"], "同花顺投资")
            self.assertNotIn("B站", set(references.account_mapping["channel"]))
            self.assertIn("B站市场部", set(references.content_hierarchy["channel"]))
            self.assertIn("B站商业化", set(references.content_hierarchy["channel"]))
            self.assertNotIn("B站", set(references.content_hierarchy["channel"]))

    def test_legacy_bilibili_account_mapping_still_matches_business_channels(self):
        legacy_mapping = pd.DataFrame(
            [
                {
                    "channel": "B站",
                    "source_account_id": "1622777305",
                    "account": "同花顺投资",
                    "mapping_source": "历史维护",
                }
            ]
        )

        lookup = account_mapping_lookup(legacy_mapping)

        self.assertEqual(lookup[("B站市场部", "1622777305")]["account"], "同花顺投资")
        self.assertEqual(lookup[("B站商业化", "1622777305")]["account"], "同花顺投资")

    def test_unknown_bilibili_mid_is_kept_for_manual_review(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            pd.DataFrame(
                [
                    {
                        "视频AVID": "av-unknown-mid",
                        "视频BVID": "bv-unknown-mid",
                        "视频标题": "B站未知账号内容",
                        "Up主mid": "999999",
                        "花费": 80,
                        "展示量": 8000,
                        "点击量": 400,
                        "应用激活数": 16,
                        "应用内付费": 4,
                    }
                ]
            ).to_csv(raw_dir / "B站.csv", index=False, encoding="utf-8-sig")

            result = analyze_input_dir(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                reference_tables_path=Path(tmp) / "config" / "reference_tables.xlsx",
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(len(result.canonical), 1)
            row = result.canonical.iloc[0]
            self.assertEqual(row["account_id"], "999999")
            self.assertEqual(row["account"], "")
            self.assertEqual(row["account_filter_status"], "")
            self.assertIn("账号映射缺失", row["review_reasons"])
            self.assertTrue(result.account_filter_details.empty)

    def test_channel_dedupe_sums_large_numeric_conflicts_and_marks_review(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            pd.DataFrame(
                [
                    {
                        "视频标题": "同一视频A",
                        "视频id": "dy-dup",
                        "素材ID": "mat-1",
                        "账号": "同花顺投资",
                        "消耗": 100,
                        "展示数": 1000,
                        "点击数": 100,
                        "激活数": 10,
                        "付费次数": 2,
                        "内容类型": "热点行情",
                    },
                    {
                        "视频标题": "同一视频A",
                        "视频id": "dy-dup",
                        "素材ID": "mat-2",
                        "账号": "同花顺投资",
                        "消耗": 120,
                        "展示数": 1005,
                        "点击数": 100,
                        "激活数": 10,
                        "付费次数": 2,
                        "内容类型": "热点行情",
                    },
                ]
            ).to_csv(raw_dir / "抖音商业化.csv", index=False, encoding="utf-8-sig")
            pd.DataFrame(
                [
                    {
                        "视频标题": "同一视频跨渠道不合并",
                        "视频id": "dy-dup",
                        "素材ID": "mat-other",
                        "账号": "同花顺投资",
                        "消耗": 70,
                        "展示数": 700,
                        "点击数": 70,
                        "激活数": 7,
                        "付费次数": 1,
                        "内容类型": "热点行情",
                    }
                ]
            ).to_csv(raw_dir / "抖音市场部.csv", index=False, encoding="utf-8-sig")

            result = analyze_input_dir(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(len(result.canonical), 2)
            commercial = result.canonical[result.canonical["channel"].eq("抖音商业化")].iloc[0]
            market = result.canonical[result.canonical["channel"].eq("抖音市场部")].iloc[0]
            self.assertEqual(commercial["merged_row_count"], 2)
            self.assertAlmostEqual(commercial["spend"], 220.0)
            self.assertAlmostEqual(commercial["impressions"], 2005.0)
            self.assertTrue(commercial["needs_manual_review"])
            self.assertIn("数值冲突", commercial["review_reasons"])
            self.assertIn("spend", commercial["conflict_details"])
            self.assertEqual(market["merged_row_count"], 1)

    def test_channel_dedupe_prefers_id_then_url_then_title_and_recomputes_rates(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "platform_group": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "dy-stable-1",
                    "content_url": "https://www.douyin.com/video/7291234567890123456",
                    "title": "同 ID 原始标题",
                    "account": "投资号",
                    "manual_category": "热点行情",
                    "spend": 100,
                    "impressions": 1000,
                    "clicks": 100,
                    "activations": 10,
                    "first_pay_count": 2,
                    "source_file": "a.csv",
                },
                {
                    "platform": "抖音",
                    "platform_group": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "dy-stable-1",
                    "content_url": "https://www.douyin.com/video/7291234567890123456?share_token=abc",
                    "title": "同 ID 补充标题 #投教",
                    "account": "财经号",
                    "manual_category": "热点行情",
                    "spend": 100,
                    "impressions": 1000,
                    "clicks": 100,
                    "activations": 10,
                    "first_pay_count": 2,
                    "source_file": "b.csv",
                },
                {
                    "platform": "抖音",
                    "platform_group": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "content_url": "https://www.douyin.com/video/7291234567890123456?foo=1",
                    "title": "链接行标题 A",
                    "account": "投资号",
                    "manual_category": "热点行情",
                    "spend": 30,
                    "impressions": 300,
                    "clicks": 30,
                    "activations": 3,
                    "first_pay_count": 1,
                    "source_file": "c.csv",
                },
                {
                    "platform": "抖音",
                    "platform_group": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "content_url": "https://www.douyin.com/video/7291234567890123456/",
                    "title": "链接行标题 B #补充",
                    "account": "财经号",
                    "manual_category": "热点行情",
                    "spend": 40,
                    "impressions": 400,
                    "clicks": 40,
                    "activations": 4,
                    "first_pay_count": 1,
                    "source_file": "d.csv",
                },
                {
                    "platform": "抖音",
                    "platform_group": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "content_url": "",
                    "title": " 标题  兜底 合并 ",
                    "account": "投资号",
                    "manual_category": "热点行情",
                    "spend": 10,
                    "impressions": 100,
                    "clicks": 10,
                    "activations": 1,
                    "first_pay_count": 0,
                    "source_file": "e.csv",
                },
                {
                    "platform": "抖音",
                    "platform_group": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "content_url": "",
                    "title": "标题兜底合并",
                    "account": "财经号",
                    "manual_category": "热点行情",
                    "spend": 15,
                    "impressions": 150,
                    "clicks": 15,
                    "activations": 2,
                    "first_pay_count": 1,
                    "source_file": "f.csv",
                },
            ]
        )

        with TemporaryDirectory() as tmp:
            result = analyze_canonical_frame(
                frame,
                "2026-04-01",
                "2026-04-27",
                category_matcher=lambda items, category_library, env_path: {},
                reference_tables_path=Path(tmp) / "reference_tables.xlsx",
            )

            self.assertEqual(len(result.canonical), 3)
            by_key = result.canonical.set_index("dedupe_key")
            id_row = by_key.loc["抖音商业化::id::7291234567890123456"]
            self.assertEqual(id_row["merged_row_count"], 4)
            self.assertEqual(id_row["title"], "同 ID 补充标题 #投教")
            self.assertAlmostEqual(id_row["spend"], 270.0)
            self.assertAlmostEqual(id_row["impressions"], 2700.0)
            self.assertAlmostEqual(id_row["activation_cost"], 10.0)
            self.assertAlmostEqual(id_row["first_pay_rate"], 6.0 / 27.0)
            self.assertIn("同ID标题不一致", id_row["review_reasons"])
            self.assertIn("title", id_row["conflict_details"])

            title_row = by_key.loc["抖音商业化::title_account::投资号::标题兜底合并"]
            other_title_row = by_key.loc["抖音商业化::title_account::财经号::标题兜底合并"]
            self.assertEqual(title_row["merged_row_count"], 1)
            self.assertEqual(other_title_row["merged_row_count"], 1)
            self.assertAlmostEqual(title_row["spend"] + other_title_row["spend"], 25.0)

    def test_generic_excel_is_treated_as_channel_from_file_name(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "知乎投放.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "标题": "知乎财经问答",
                            "内容ID": "zh-1",
                            "账号": "同花顺投资",
                            "内容类型": "投教问答",
                            "消耗": 30,
                            "展示数": 300,
                            "点击数": 30,
                            "激活数": 3,
                            "付费次数": 1,
                        }
                    ]
                ).to_excel(writer, sheet_name="投放数据", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                category_matcher=lambda items, category_library, env_path: {},
            )

            row = result.canonical.iloc[0]
            self.assertEqual(row["channel"], "知乎投放")
            self.assertEqual(row["account"], "同花顺投资")
            self.assertEqual(row["content_id"], "zh-1")
            self.assertEqual(row["category_l2"], "")
            self.assertEqual(row["analysis_status"], "不可分析")
            self.assertEqual(row["unanalyzable_reason"], "平台不在复盘范围")

    def test_market_and_new_platform_sources_are_mapped_to_distinct_channels(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "小红书（市场部）.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "笔记/素材ID": "xhs-market-1",
                            "笔记/素材链接": "https://xhs.example/1",
                            "账号": "问财",
                            "消费": 10,
                            "展现量": 100,
                            "点击量": 20,
                            "激活数(转化时间)": 3,
                            "首次付费次数(转化时间)": 1,
                        }
                    ]
                ).to_excel(writer, sheet_name="计划-数据", index=False)
            with pd.ExcelWriter(raw_dir / "微信市场部.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "创意名称": "微信创意一",
                            "花费": 20,
                            "曝光次数": 200,
                            "点击次数": 30,
                            "APP激活次数": 4,
                            "注册次数": 2,
                        }
                    ]
                ).to_excel(writer, sheet_name="Sheet1", index=False)
            with pd.ExcelWriter(raw_dir / "腾讯（市场部）.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "链接": "https://tencent.example/creative",
                            "花费": 30,
                            "曝光次数": 300,
                            "点击次数": 40,
                            "APP激活次数": 5,
                            "注册次数（点击归因）": 2,
                        }
                    ]
                ).to_excel(writer, sheet_name="Sheet1", index=False)
            with pd.ExcelWriter(raw_dir / "视频号投放.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "创意名称": "视频号创意一",
                            "花费": 25,
                            "曝光次数": 250,
                            "点击次数": 35,
                            "APP激活次数": 7,
                            "注册次数": 3,
                        }
                    ]
                ).to_excel(writer, sheet_name="Sheet1", index=False)
            with pd.ExcelWriter(raw_dir / "抖音原生-达人数据情况（商业化）.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "视频链接": "https://douyin.example/video/1",
                            "账号": "同花顺投资",
                            "消耗": 40,
                            "展示数": 400,
                            "激活数": 6,
                            "付费数": 3,
                        }
                    ]
                ).to_excel(writer, sheet_name="Sheet4", index=False)

            result = analyze_input_dir(
                raw_dir,
                "2026-05-08",
                "2026-05-14",
                category_matcher=lambda items, category_library, env_path: {},
            )

            channels = set(result.canonical["channel"])
            self.assertIn("小红书市场部", channels)
            self.assertIn("微信市场部", channels)
            self.assertIn("抖音商业化", channels)
            self.assertNotIn("微信/腾讯/视频号市场部", channels)
            self.assertNotIn("腾讯市场部", channels)
            self.assertNotIn("达人数据", channels)
            self.assertNotIn("抖音达人内容", channels)
            social_rows = result.canonical[result.canonical["channel"].eq("微信市场部")]
            self.assertEqual(len(social_rows), 3)
            self.assertEqual(set(social_rows["platform"]), {"微信"})
            self.assertEqual(set(social_rows["platform_group"]), {"微信"})

    def test_core_workbook_uses_fixed_three_sheet_schema(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "知乎投放.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "标题": "知乎财经问答",
                            "内容ID": "zh-1",
                            "账号": "同花顺投资",
                            "内容类型": "投教问答",
                            "消耗": 30,
                            "展示数": 300,
                            "点击数": 30,
                            "激活数": 3,
                            "付费次数": 1,
                            "自定义列": "保留原始字段",
                        }
                    ]
                ).to_excel(writer, sheet_name="投放数据", index=False)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            self.assertTrue(result.core_recap_xlsx.exists())
            workbook = load_workbook(result.core_recap_xlsx, read_only=True)
            self.assertEqual(workbook.sheetnames, CORE_ANALYSIS_SHEETS)
            headers = [cell.value for cell in next(workbook["清洗后素材表"].iter_rows(max_row=1))]
            self.assertEqual(
                headers,
                [
                    "周期",
                    "平台",
                    "渠道",
                    "原始标题",
                    "标准标题",
                    "作品链接",
                    "作品ID/BV号",
                    "巨量素材ID",
                    "消耗",
                    "曝光",
                    "飞书匹配结果",
                    "飞书匹配标题",
                    "内容类型",
                    "是否可分析",
                    "不可分析原因",
                    "来源文件",
                    "来源行号",
                ],
            )
            workbook.close()

    def test_bilibili_numeric_mid_is_normalized_without_decimal_suffix(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            raw_dir.mkdir()
            pd.DataFrame(
                [
                    {
                        "视频AVID": "av-mid",
                        "视频BVID": "bv-mid",
                        "视频标题": "B站财经内容",
                        "Up主mid": 1622777305.0,
                        "花费": 80,
                        "展示量": 8000,
                        "点击量": 400,
                        "应用激活数": 16,
                        "应用内付费": 4,
                    }
                ]
            ).to_csv(raw_dir / "B站.csv", index=False, encoding="utf-8-sig")

            result = analyze_input_dir(
                raw_dir,
                "2026-04-01",
                "2026-04-27",
                category_matcher=lambda items, category_library, env_path: {},
            )

            self.assertEqual(result.canonical.iloc[0]["account_id"], "1622777305")
            self.assertEqual(result.canonical.iloc[0]["account"], "同花顺投资")

    def test_workflow_builds_data_quality_report_and_core_workbook(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "B站.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "视频AVID": "av-unknown",
                            "视频BVID": "bv-unknown",
                            "视频标题": "一个暂时无法从标题判断的内容",
                            "Up主mid": "1622777305",
                            "花费": 100.0,
                            "展示量": 0,
                            "点击量": 10,
                            "应用激活数": 1,
                            "应用内付费": 0,
                        }
                    ]
                ).to_excel(writer, sheet_name="sheet1", index=False)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            quality = result.data_quality.set_index("metric")
            self.assertNotIn("一级分类缺失率", quality.index)
            self.assertIn("二级分类缺失率", quality.index)
            self.assertEqual(quality.loc["展示为0但点击大于0", "count"], 1)
            self.assertTrue(result.core_recap_xlsx.exists())

    def test_workflow_summarizes_channels_without_external_totals(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_feishu_ledger_fixture()):
                result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            self.assertEqual(set(result.channel_summary["channel"]), {"B站市场部", "小红书商业化", "抖音商业化", "抖音市场部"})
            channel_spend = result.channel_summary.set_index("channel")["spend"]
            self.assertAlmostEqual(float(channel_spend.loc["B站市场部"]), 100.0)
            self.assertAlmostEqual(float(channel_spend.loc["小红书商业化"]), 100.0)
            summary = result.unanalyzable_summary.set_index("渠道")
            self.assertAlmostEqual(summary.loc["B站市场部", "不可分析消耗"], 100.0)
            self.assertAlmostEqual(summary.loc["小红书商业化", "不可分析消耗"], 100.0)
            self.assertAlmostEqual(summary.loc["抖音商业化", "不可分析消耗"], 0.0)

            self.assertNotIn("spend_ratio", result.canonical.columns)
            self.assertNotIn("spend_calibrated", result.canonical.columns)

    def test_workflow_writes_core_analysis_workbook_only(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            self.assertTrue(result.core_recap_xlsx.exists())
            self.assertEqual(result.core_recap_xlsx, output_dir / "content_recap_core.xlsx")
            workbook = load_workbook(result.core_recap_xlsx, read_only=True)
            self.assertEqual(workbook.sheetnames, CORE_ANALYSIS_SHEETS)
            workbook.close()

    def test_archived_workflow_ui_only_persists_core_tables_without_report_or_ai_calls(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)
            pd.DataFrame(
                [
                    {
                        "视频标题": "已知类型样本",
                        "视频id": "dy-known-ui-only",
                        "素材ID": "mat-known-ui-only",
                        "消耗": 10.0,
                        "内容类型": "资讯",
                    },
                    {
                        "视频标题": "完全陌生主题",
                        "视频id": "dy-pending-ui-only",
                        "素材ID": "mat-pending-ui-only",
                        "消耗": 9.0,
                        "内容类型": "",
                    },
                ]
            ).to_csv(raw_dir / "抖音市场部.csv", index=False, encoding="utf-8-sig")
            env_path = tmp_path / ".env"
            env_path.write_text(
                "DEEPSEEK_API_KEY=test-key\nDEEPSEEK_BASE_URL=https://api.deepseek.com\n",
                encoding="utf-8",
            )
            reviewed_batches = []

            def category_matcher(items, category_library, env_path_arg):
                reviewed_batches.append((len(items), tuple(category_library)))
                return {}

            with patch(
                "ops_data_workflow.topic_analysis.group_topic_labels",
                side_effect=AssertionError("DeepSeek topic call"),
            ):
                result = run_archived_workflow(
                    raw_dir,
                    "2026-04-01",
                    "2026-04-27",
                    output_root=tmp_path / "outputs",
                    processed_root=tmp_path / "processed",
                    db_path=tmp_path / ".runtime" / "workflow.sqlite3",
                    env_path=env_path,
                    category_matcher=category_matcher,
                    output_mode="ui_only",
                    enable_deepseek=True,
                    enable_external_context=False,
                )

            self.assertFalse(reviewed_batches)
            self.assertTrue((result.archive_dir / "cleaned.xlsx").exists())
            self.assertTrue(result.core_recap_xlsx.exists())
            self.assertEqual(result.core_recap_xlsx, result.archive_dir / "content_recap_core.xlsx")
            self.assertFalse((tmp_path / "outputs" / result.batch_id).exists())
            self.assertEqual(result.ai_summary, "")
            with closing(sqlite3.connect(tmp_path / ".runtime" / "workflow.sqlite3")) as conn:
                canonical_count = conn.execute(
                    "select count(*) from canonical_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                topic_providers = [
                    row[0]
                    for row in conn.execute(
                        "select distinct provider from topic_label_items where batch_id = ?",
                        (result.batch_id,),
                    ).fetchall()
                ]
                ai_count = conn.execute(
                    "select count(*) from ai_reports where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                asset_count = conn.execute(
                    "select count(*) from cleaned_asset_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                summary_count = conn.execute(
                    "select count(*) from unanalyzable_summary_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
            self.assertGreater(canonical_count, 0)
            self.assertEqual(asset_count, canonical_count)
            self.assertGreater(summary_count, 0)
            self.assertNotIn("deepseek", topic_providers)
            self.assertEqual(ai_count, 0)

    def test_workflow_writes_core_recap_when_category_spend_is_blank(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "抖音商业化.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "视频标题": "有曝光但无消耗的内容",
                            "视频id": "dy-organic",
                            "视频链接": "https://douyin.example/organic",
                            "账号": "同花顺投资",
                            "内容类型": "股友说",
                            "消耗": "",
                            "展示数": 5317,
                            "激活数": "",
                            "付费次数": "",
                        }
                    ]
                ).to_excel(writer, sheet_name="Sheet2", index=False)

            result = run_workflow(raw_dir, "2026-04-10", "2026-04-16", output_dir)

            self.assertTrue(result.core_recap_xlsx.exists())
            self.assertEqual(list(result.canonical["content_id"]), [""])
            self.assertEqual(list(result.canonical["material_id"]), ["dy-organic"])
            self.assertEqual(list(result.canonical["ad_material_id"]), ["dy-organic"])

    def test_core_tables_use_readable_chinese_column_names(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            for headers in [
                list(result.cleaned_asset_table.columns),
                list(result.content_recap_table.columns),
                list(result.unanalyzable_summary.columns),
            ]:
                self.assertFalse(
                    any(_looks_like_english_field_name(header) for header in headers if header),
                    headers,
                )
            self.assertIn("作品ID/BV号", result.cleaned_asset_table.columns)
            self.assertIn("不可分析原因", result.cleaned_asset_table.columns)
            self.assertIn("不可分析消耗占比", result.unanalyzable_summary.columns)

    def test_workflow_accepts_only_the_four_raw_platform_excels(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)
            _remove_total_fixture(raw_dir)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            self.assertEqual(len(result.canonical), 6)
            self.assertTrue(result.core_recap_xlsx.exists())
            workbook = load_workbook(result.core_recap_xlsx, read_only=True)
            self.assertEqual(workbook.sheetnames, CORE_ANALYSIS_SHEETS)
            workbook.close()

    def test_missing_values_keep_raw_blanks_without_forcing_unmatched_categories(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)
            canonical = result.canonical

            bilibili = canonical[canonical["content_id"].eq("bv1")].iloc[0]
            douyin_market = canonical[canonical["material_id"].eq("mat-3")].iloc[0]
            xhs_missing_secondary = canonical[canonical["content_id"].eq("note-2")].iloc[0]

            self.assertEqual(bilibili["content_category"], "")
            self.assertEqual(bilibili["analysis_status"], "不可分析")
            self.assertTrue(pd.isna(douyin_market["clicks"]))
            self.assertEqual(xhs_missing_secondary["content_category"], "")
            self.assertEqual(xhs_missing_secondary["analysis_status"], "不可分析")

    def test_workflow_collects_raw_category_statistics(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            stats = result.raw_category_stats
            self.assertIn("热点行情", set(stats["value"]))
            self.assertIn("图文", set(stats["value"]))
            self.assertIn("股友说", set(stats["value"]))

    def test_workflow_writes_unanalyzable_summary_instead_of_total_summary_workbook(self):
        with TemporaryDirectory() as tmp:
            raw_dir = Path(tmp) / "raw"
            output_dir = Path(tmp) / "out"
            raw_dir.mkdir()
            _write_raw_fixture(raw_dir)

            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_feishu_ledger_fixture()):
                result = run_workflow(raw_dir, "2026-04-01", "2026-04-27", output_dir)

            self.assertFalse(result.unanalyzable_summary.empty)
            self.assertIn("不可分析素材占比", result.unanalyzable_summary.columns)
            self.assertEqual(int(result.unanalyzable_summary["可分析素材数"].sum()), 3)

    def test_archived_workflow_persists_safe_public_metadata_enrichment_from_cache(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            metadata_cache = tmp_path / "metadata-cache"
            cache_item = metadata_cache / "bilibili" / "BV1workflow1.json"
            cache_item.parent.mkdir(parents=True)
            cache_item.write_text(
                json.dumps(
                    {
                        "id": "BV1workflow1",
                        "link": "https://www.bilibili.com/video/BV1workflow1/",
                        "title": "缓存标题",
                        "tags": "财经",
                        "published_at": "2026-05-09",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            pd.DataFrame(
                [
                    {
                        "视频BVID": "BV1workflow1",
                        "花费": 10.0,
                        "展示量": 100,
                        "点击量": 10,
                        "应用激活数": 2,
                        "应用内首次付费次数": 1,
                    }
                ]
            ).to_csv(raw_dir / "B站数据.csv", index=False, encoding="utf-8-sig")

            result = run_archived_workflow(
                raw_dir,
                "2026-05-08",
                "2026-05-14",
                output_root=tmp_path / "outputs",
                processed_root=tmp_path / "processed",
                db_path=tmp_path / ".runtime" / "workflow.sqlite3",
                output_mode="ui_only",
                enable_deepseek=False,
                enable_external_context=False,
                metadata_enrichment_mode="safe_public",
                metadata_cache_dir=metadata_cache,
            )

            with closing(sqlite3.connect(tmp_path / ".runtime" / "workflow.sqlite3")) as conn:
                row = conn.execute(
                    """
                    select content_url, title, source_time, metadata_source, metadata_tags
                    from canonical_items
                    where batch_id = ?
                    """,
                    (result.batch_id,),
                ).fetchone()

            self.assertEqual(row[0], "https://www.bilibili.com/video/BV1workflow1/")
            self.assertEqual(row[1], "缓存标题")
            self.assertEqual(row[2], "2026-05-09")
            self.assertEqual(row[3], "metadata_cache")
            self.assertEqual(row[4], "财经")

    def test_archived_workflow_can_force_reclean_existing_cleaned_workbook_for_metadata_enrichment(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            stale_cleaned = raw_dir / "cleaned.xlsx"
            with pd.ExcelWriter(stale_cleaned, engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "platform": "B站",
                            "platform_group": "B站",
                            "channel": "B站",
                            "content_id": "BV1stale001",
                            "title": "旧清洗标题",
                            "content_url": "",
                            "spend": 1.0,
                            "impressions": 1.0,
                            "clicks": 1.0,
                            "activations": 1.0,
                            "first_pay_count": 1.0,
                        }
                    ]
                ).to_excel(writer, sheet_name="清洗后明细", index=False)
            pd.DataFrame(
                [
                    {
                        "视频BVID": "BV1fresh001",
                        "花费": 10.0,
                        "展示量": 100,
                        "点击量": 10,
                        "应用激活数": 2,
                        "应用内首次付费次数": 1,
                    }
                ]
            ).to_csv(raw_dir / "B站数据.csv", index=False, encoding="utf-8-sig")

            metadata_cache = tmp_path / "metadata-cache"
            cache_item = metadata_cache / "bilibili" / "BV1fresh001.json"
            cache_item.parent.mkdir(parents=True)
            cache_item.write_text(
                json.dumps(
                    {
                        "id": "BV1fresh001",
                        "link": "https://www.bilibili.com/video/BV1fresh001/",
                        "title": "原始表补全标题",
                        "published_at": "2026-05-09",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = run_archived_workflow(
                raw_dir,
                "2026-05-08",
                "2026-05-14",
                output_root=tmp_path / "outputs",
                processed_root=tmp_path / "processed",
                db_path=tmp_path / ".runtime" / "workflow.sqlite3",
                output_mode="ui_only",
                enable_deepseek=False,
                enable_external_context=False,
                metadata_enrichment_mode="safe_public",
                metadata_cache_dir=metadata_cache,
                force_reclean=True,
            )

            self.assertIn("BV1fresh001", set(result.canonical["content_id"]))
            self.assertNotIn("BV1stale001", set(result.canonical["content_id"]))
            row = result.canonical[result.canonical["content_id"].eq("BV1fresh001")].iloc[0]
            self.assertEqual(row["title"], "原始表补全标题")
            self.assertEqual(row["metadata_source"], "metadata_cache")
            self.assertIn("BV1fresh001", set(result.top_content_items["content_id"]))
            self.assertEqual(
                result.top_content_items[result.top_content_items["content_id"].eq("BV1fresh001")].iloc[0]["title"],
                "原始表补全标题",
            )
            self.assertEqual(result.cleaned_asset_table.iloc[0]["标准标题"], "原始表补全标题")

    def test_refresh_historical_source_periods_rebuilds_from_raw_sources_with_safe_public_metadata(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            data_root = tmp_path / "data"
            source_dir = data_root / "weeks" / "20260508-20260514"
            source_dir.mkdir(parents=True)
            pd.DataFrame(
                [
                    {
                        "视频BVID": "BV1history1",
                        "花费": 20.0,
                        "展示量": 200,
                        "点击量": 20,
                        "应用激活数": 4,
                        "应用内首次付费次数": 2,
                    }
                ]
            ).to_csv(source_dir / "B站数据.csv", index=False, encoding="utf-8-sig")
            processed_dir = tmp_path / "processed" / "20260508-20260514" / "upload:week:20260508-20260514"
            processed_dir.mkdir(parents=True)
            with pd.ExcelWriter(processed_dir / "cleaned.xlsx", engine="openpyxl") as writer:
                pd.DataFrame([{"channel": "B站", "content_id": "BV1oldhist"}]).to_excel(
                    writer,
                    sheet_name="清洗后明细",
                    index=False,
                )
            metadata_cache = tmp_path / "metadata-cache"
            cache_item = metadata_cache / "bilibili" / "BV1history1.json"
            cache_item.parent.mkdir(parents=True)
            cache_item.write_text(
                json.dumps(
                    {
                        "id": "BV1history1",
                        "link": "https://www.bilibili.com/video/BV1history1/",
                        "title": "历史重算标题",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            results = refresh_historical_source_periods(
                data_root=data_root,
                processed_root=tmp_path / "processed",
                output_root=tmp_path / "outputs",
                db_path=tmp_path / ".runtime" / "workflow.sqlite3",
                metadata_cache_dir=metadata_cache,
            )

            self.assertEqual([result.batch_id for result in results], ["upload:week:20260508-20260514"])
            with closing(sqlite3.connect(tmp_path / ".runtime" / "workflow.sqlite3")) as conn:
                canonical_row = conn.execute(
                    """
                    select content_id, title, metadata_source
                    from canonical_items
                    where batch_id = 'upload:week:20260508-20260514'
                    """
                ).fetchone()
                asset_row = conn.execute(
                    """
                    select "作品ID/BV号", "标准标题", "作品链接"
                    from cleaned_asset_items
                    where batch_id = 'upload:week:20260508-20260514'
                    """
                ).fetchone()

            self.assertEqual(canonical_row, ("BV1history1", "历史重算标题", "metadata_cache"))
            self.assertEqual(asset_row, ("BV1history1", "历史重算标题", "https://www.bilibili.com/video/BV1history1/"))


if __name__ == "__main__":
    unittest.main()
