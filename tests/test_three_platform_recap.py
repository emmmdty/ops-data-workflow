from pathlib import Path
from tempfile import TemporaryDirectory
import sqlite3
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

from ops_data_workflow.analysis_scope import apply_analysis_scope
from ops_data_workflow.asset_matching import match_assets_to_ledger
from ops_data_workflow.exports import ATTRIBUTION_RECAP_SHEETS, CORE_RECAP_SHEETS, CORE_RECAP_WORKBOOK, write_core_recap_workbook
from ops_data_workflow.platform_normalizers import normalize_platform_identities
from ops_data_workflow.recap_dataset import (
    CLEANED_ASSET_COLUMNS,
    CONTENT_RECAP_COLUMNS,
    UNANALYZABLE_SUMMARY_COLUMNS,
    build_cleaned_asset_table,
    build_content_recap_table,
    build_unanalyzable_summary,
)
from ops_data_workflow.workflow import run_archived_workflow


def _ledger() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "platform": "小红书",
                "content_id": "65f00000abcdef",
                "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                "title": "小红书自有标题",
                "account": "投资号",
                "content_type": "资讯",
                "source_file": "harvester_feishu",
                "source_sheet": "小红书",
                "source_row": 2,
                "title_key": "",
                "title_key_no_tags": "",
            },
            {
                "platform": "B站",
                "content_id": "BV1abcde2345",
                "content_url": "https://www.bilibili.com/video/BV1abcde2345/",
                "title": "B站自有标题",
                "account": "投资号",
                "content_type": "长视频",
                "source_file": "harvester_feishu",
                "source_sheet": "B站",
                "source_row": 3,
                "title_key": "",
                "title_key_no_tags": "",
            },
            {
                "platform": "抖音",
                "content_id": "",
                "content_url": "https://www.douyin.com/video/7291234567890123456",
                "title": "抖音真实标题",
                "account": "投资号",
                "content_type": "股友说",
                "source_file": "harvester_feishu",
                "source_sheet": "抖音",
                "source_row": 4,
                "title_key": "",
                "title_key_no_tags": "抖音真实标题",
            },
        ]
    )


CORE_ANALYSIS_SHEETS = CORE_RECAP_SHEETS + ATTRIBUTION_RECAP_SHEETS


class ThreePlatformRecapTests(unittest.TestCase):
    def test_platform_normalizers_extract_work_identity_and_demote_douyin_ad_id(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "小红书",
                    "channel": "小红书商业化",
                    "content_id": "",
                    "material_id": "",
                    "title": "小红书投放标题",
                    "content_url": "https://www.xiaohongshu.com/discovery/item/65f00000abcdef?xsec_token=abc",
                },
                {
                    "platform": "B站",
                    "channel": "B站市场部",
                    "content_id": "",
                    "material_id": "单元-BV1abcde2345",
                    "title": "B站投放标题",
                    "content_url": "",
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "v02033g10000giant",
                    "material_id": "7391234567890123456",
                    "title": "A组-推送素材-0420",
                    "content_url": "",
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "material_id": "mat-1",
                    "title": "抖音真实标题 #投资",
                    "content_url": "https://www.douyin.com/video/7291234567890123456",
                },
            ]
        )

        normalized = normalize_platform_identities(frame)

        xhs, bili, dy_ad_id_only, dy_url = [normalized.iloc[index] for index in range(4)]
        self.assertEqual(xhs["work_id"], "65f00000abcdef")
        self.assertEqual(xhs["work_url"], "https://www.xiaohongshu.com/explore/65f00000abcdef")
        self.assertEqual(bili["work_id"], "BV1abcde2345")
        self.assertEqual(bili["work_url"], "https://www.bilibili.com/video/BV1abcde2345/")
        self.assertEqual(dy_ad_id_only["work_id"], "")
        self.assertEqual(dy_ad_id_only["ad_material_id"], "7391234567890123456")
        self.assertIn("非真实作品标题", dy_ad_id_only["normalization_reason"])
        self.assertEqual(dy_url["work_id"], "7291234567890123456")
        self.assertEqual(dy_url["ad_material_id"], "mat-1")
        self.assertEqual(dy_url["standard_title"], "抖音真实标题")

    def test_platform_normalizer_does_not_copy_douyin_work_id_to_material_id(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音市场部",
                    "content_id": "7632278691925609769",
                    "material_id": "",
                    "title": "存多少钱才可以提前退休",
                    "content_url": "https://www.douyin.com/video/7632278691925609769",
                }
            ]
        )

        normalized = normalize_platform_identities(frame)
        row = normalized.iloc[0]

        self.assertEqual(row["content_id"], "7632278691925609769")
        self.assertEqual(row["work_id"], "7632278691925609769")
        self.assertEqual(row["material_id"], "")
        self.assertEqual(row["ad_material_id"], "")

    def test_platform_normalizer_standardizes_douyin_copied_share_text(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "material_id": "mat-1",
                    "title": "1.28 tRk:/ 人和人的缘分就像炒股 #同花顺股友说 #投资 https://www.douyin.com/video/7291234567890123456?previous_page=web_code_link 复制此链接，打开Dou音搜索，直接观看视频！",
                    "content_url": "",
                }
            ]
        )

        normalized = normalize_platform_identities(frame)
        row = normalized.iloc[0]

        self.assertEqual(row["work_id"], "7291234567890123456")
        self.assertEqual(row["content_id"], "7291234567890123456")
        self.assertEqual(row["work_url"], "https://www.douyin.com/video/7291234567890123456")
        self.assertEqual(row["standard_title"], "人和人的缘分就像炒股")
        self.assertEqual(row["normalization_status"], "ok")

    def test_platform_normalizer_extracts_douyin_note_url_without_promoting_plain_ad_id(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音市场部",
                    "content_id": "7641566631126256938",
                    "material_id": "",
                    "title": "裸ID标题",
                    "content_url": "",
                },
                {
                    "platform": "抖音",
                    "channel": "抖音市场部",
                    "content_id": "",
                    "material_id": "",
                    "title": "图文标题",
                    "content_url": "https://www.douyin.com/note/7641564223058824491",
                },
            ]
        )

        normalized = normalize_platform_identities(frame)

        self.assertEqual(list(normalized["work_id"]), ["", "7641564223058824491"])
        self.assertEqual(
            list(normalized["work_url"]),
            [
                "",
                "https://www.douyin.com/video/7641564223058824491",
            ],
        )

    def test_platform_normalizer_keeps_douyin_shortlink_share_pending_resolution(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "material_id": "mat-1",
                    "title": "5.66 abc:/ 短链标题 #财经 https://v.douyin.com/AbCdEf/ 复制此链接，打开抖音搜索，直接观看视频！",
                    "content_url": "",
                }
            ]
        )

        normalized = normalize_platform_identities(frame)
        row = normalized.iloc[0]

        self.assertEqual(row["work_id"], "")
        self.assertEqual(row["content_id"], "")
        self.assertEqual(row["work_url"], "https://v.douyin.com/AbCdEf")
        self.assertEqual(row["standard_title"], "短链标题")
        self.assertEqual(row["normalization_status"], "pending_enrichment")
        self.assertIn("抖音URL解析失败", row["normalization_reason"])

    def test_platform_normalizer_separates_douyin_work_and_ad_material_links(self):
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "",
                    "material_id": "7631242926141521926",
                    "title": "只有巨量素材链接",
                    "content_url": "",
                    "ad_material_url": "https://巨量.example/video.mp4",
                    "ad_cover_url": "https://巨量.example/cover.jpg",
                }
            ]
        )

        normalized = normalize_platform_identities(frame)
        row = normalized.iloc[0]

        self.assertEqual(row["work_id"], "")
        self.assertEqual(row["work_url"], "")
        self.assertEqual(row["content_id"], "")
        self.assertEqual(row["content_url"], "")
        self.assertEqual(row["ad_material_id"], "7631242926141521926")
        self.assertEqual(row["ad_material_url"], "https://巨量.example/video.mp4")
        self.assertEqual(row["ad_cover_url"], "https://巨量.example/cover.jpg")

    def test_asset_matching_uses_platform_identity_and_never_douyin_ad_id(self):
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "小红书",
                        "channel": "小红书商业化",
                        "content_id": "",
                        "material_id": "",
                        "title": "投放标题",
                        "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef?xsec=1",
                    },
                    {
                        "platform": "B站",
                        "channel": "B站市场部",
                        "content_id": "",
                        "material_id": "BV1abcde2345",
                        "title": "投放标题",
                        "content_url": "",
                    },
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "content_id": "7291234567890123456",
                        "material_id": "巨量素材-1",
                        "title": "投放侧改写标题",
                        "content_url": "",
                    },
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "content_id": "",
                        "material_id": "",
                        "title": "抖音真实标题 #财经",
                        "content_url": "",
                    },
                ]
            )
        )

        matched = match_assets_to_ledger(frame, _ledger())

        self.assertEqual(list(matched["match_status"]), ["已匹配", "已匹配", "未匹配", "已匹配"])
        self.assertEqual(matched.iloc[0]["match_source"], "作品ID")
        self.assertEqual(matched.iloc[1]["match_source"], "BV号")
        self.assertEqual(matched.iloc[2]["match_reason"], "未匹配飞书自有内容")
        self.assertEqual(matched.iloc[3]["match_source"], "标准标题")
        self.assertEqual(matched.iloc[3]["matched_content_type"], "股友说")

    def test_asset_matching_does_not_match_douyin_plain_material_id_as_work_id(self):
        ledger = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "content_id": "7626286546770968627",
                    "content_url": "https://www.douyin.com/video/7626286546770968627",
                    "title": "台账真实作品",
                    "account": "投资号",
                    "content_type": "资讯",
                    "title_key": "",
                    "title_key_no_tags": "",
                },
                {
                    "platform": "抖音",
                    "content_id": "7594830477777751338",
                    "content_url": "https://www.douyin.com/video/7594830477777751338",
                    "title": "高价值真实选题",
                    "account": "投资号",
                    "content_type": "股友说",
                    "category_l1": "股友说",
                    "category_l2": "股民教学",
                    "title_key": "",
                    "title_key_no_tags": "",
                },
            ]
        )
        frame = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "7626286546770968627",
                    "material_id": "7626286546770968627",
                    "work_id": "",
                    "work_url": "",
                    "title": "投放侧改写标题",
                    "content_url": "",
                },
                {
                    "platform": "抖音",
                    "channel": "抖音商业化",
                    "content_id": "7626286546770968627",
                    "material_id": "7626286546770968627",
                    "work_id": "",
                    "work_url": "",
                    "title": "高价值真实选题 #财经",
                    "content_url": "",
                },
            ]
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(matched.iloc[0]["match_status"], "未匹配")
        self.assertEqual(matched.iloc[0]["match_reason"], "未匹配飞书自有内容")
        self.assertEqual(matched.iloc[1]["match_status"], "已匹配")
        self.assertEqual(matched.iloc[1]["match_source"], "标准标题")
        self.assertEqual(matched.iloc[1]["work_id"], "7594830477777751338")
        self.assertEqual(matched.iloc[1]["work_url"], "https://www.douyin.com/video/7594830477777751338")
        self.assertEqual(matched.iloc[1]["content_id"], "7594830477777751338")

    def test_asset_matching_allows_douyin_title_prefix_and_punctuation_drift(self):
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "为什么开始炒股后物欲变低了? #财经 #同花顺投资",
                        "material_id": "7631242926141521926",
                    }
                ]
            )
        )
        ledger = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "content_id": "7631144938532539711",
                    "content_url": "https://www.douyin.com/video/7631144938532539711",
                    "title": "01/03 为什么开始炒股后物欲变低了？",
                    "account": "投资号",
                    "content_type": "资讯",
                    "title_key": "",
                    "title_key_no_tags": "",
                }
            ]
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(matched.iloc[0]["match_status"], "已匹配")
        self.assertEqual(matched.iloc[0]["match_source"], "标准标题")
        self.assertEqual(matched.iloc[0]["matched_content_type"], "资讯")
        self.assertEqual(matched.iloc[0]["matched_account"], "投资号")

    def test_asset_matching_keeps_douyin_welfare_tag_separate_from_other_accounts(self):
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "你是家族里第一个接触股市的人 #同花顺福利官 #财经 #投资理财",
                        "material_id": "7631402270895571007",
                    }
                ]
            )
        )
        ledger = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "content_id": "7638887119435468084",
                    "content_url": "https://www.douyin.com/video/7638887119435468084",
                    "title": "你是家族里第一个接触股市的人",
                    "tags": "#财经 #同花顺股友说 #同花顺社区",
                    "account": "股民社区",
                    "content_type": "股友说",
                    "title_key": "",
                    "title_key_no_tags": "",
                }
            ]
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(matched.iloc[0]["match_status"], "未匹配")
        self.assertEqual(matched.iloc[0]["match_reason"], "未匹配飞书自有内容")

    def test_asset_matching_allows_douyin_title_and_tag_mixed_text(self):
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "#大数据 扫H，你怕了吗？#财经 #涨知识",
                        "material_id": "7249502380530040835",
                    },
                    {
                        "platform": "抖音",
                        "channel": "抖音商业化",
                        "title": "换手率#口诀来了，怕你们记不住，写成了歌#财经 #改编歌曲",
                        "material_id": "7249502467044540419",
                    },
                ]
            )
        )
        ledger = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "content_id": "7033371590762990862",
                    "content_url": "https://www.douyin.com/video/7033371590762990862",
                    "title": "扫H，你怕了吗？",
                    "tags": "#大数据 #财经 #涨知识",
                    "account": "投资号",
                    "content_type": "无",
                    "title_key": "",
                    "title_key_no_tags": "",
                },
                {
                    "platform": "抖音",
                    "content_id": "7031042784765218078",
                    "content_url": "https://www.douyin.com/video/7031042784765218078",
                    "title": "换手率，怕你们记不住，写成了歌",
                    "tags": "#口诀 #财经 #改编歌曲 #口诀来了",
                    "account": "投资号",
                    "content_type": "无",
                    "title_key": "",
                    "title_key_no_tags": "",
                },
            ]
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(list(matched["match_status"]), ["已匹配", "已匹配"])
        self.assertEqual(list(matched["match_source"]), ["标准标题", "标准标题"])
        self.assertEqual(list(matched["matched_ledger_title"]), ["扫H，你怕了吗？", "换手率，怕你们记不住，写成了歌"])

    def test_asset_matching_prefers_douyin_work_id_before_title(self):
        ledger = pd.DataFrame(
            [
                {
                    "platform": "抖音",
                    "content_id": "7641566631126256938",
                    "content_url": "https://www.douyin.com/note/7641566631126256938",
                    "title": "国内核心供应商一览 猎鹰火箭里的 “中国制造”，到底有多关键？",
                    "account": "投资号",
                    "content_type": "行业品种产业链解析",
                    "category_l1": "盘点",
                    "category_l2": "行业品种产业链解析",
                    "title_key": "",
                    "title_key_no_tags": "",
                }
            ]
        )
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "抖音",
                        "channel": "抖音市场部",
                        "content_id": "",
                        "material_id": "",
                        "content_url": "https://www.douyin.com/video/7641566631126256938",
                        "title": "SpaceX 国内核心供应商一览 猎鹰火箭里的 “中国制造”",
                    }
                ]
            )
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(matched.iloc[0]["match_status"], "已匹配")
        self.assertEqual(matched.iloc[0]["match_source"], "作品ID")
        self.assertEqual(matched.iloc[0]["match_key"], "7641566631126256938")
        self.assertEqual(matched.iloc[0]["category_l1"], "盘点")
        self.assertEqual(matched.iloc[0]["category_l2"], "行业品种产业链解析")

    def test_asset_matching_prefers_platform_work_id_when_id_and_url_conflict(self):
        ledger = pd.DataFrame(
            [
                {
                    "platform": "小红书",
                    "content_id": "65f00000abcdef",
                    "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                    "title": "错误ID标题",
                    "content_type": "错误类型",
                },
                {
                    "platform": "小红书",
                    "content_id": "65f00000fedcba",
                    "content_url": "https://www.xiaohongshu.com/explore/65f00000fedcba",
                    "title": "链接标题",
                    "content_type": "链接类型",
                },
                {
                    "platform": "B站",
                    "content_id": "BV1111111111",
                    "content_url": "https://www.bilibili.com/video/BV1111111111/",
                    "title": "错误BV标题",
                    "content_type": "错误BV类型",
                },
                {
                    "platform": "B站",
                    "content_id": "BV2222222222",
                    "content_url": "https://www.bilibili.com/video/BV2222222222/",
                    "title": "链接BV标题",
                    "content_type": "链接BV类型",
                },
            ]
        )
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "小红书",
                        "channel": "小红书商业化",
                        "material_id": "65f00000abcdef",
                        "content_url": "https://www.xiaohongshu.com/explore/65f00000fedcba",
                        "title": "投放标题",
                    },
                    {
                        "platform": "B站",
                        "channel": "B站市场部",
                        "material_id": "BV1111111111",
                        "content_url": "https://www.bilibili.com/video/BV2222222222/",
                        "title": "投放标题",
                    },
                ]
            )
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(matched.iloc[0]["match_source"], "作品ID")
        self.assertEqual(matched.iloc[0]["matched_ledger_title"], "错误ID标题")
        self.assertEqual(matched.iloc[1]["match_source"], "BV号")
        self.assertEqual(matched.iloc[1]["matched_ledger_title"], "链接BV标题")

    def test_bilibili_matching_replaces_bv_placeholder_title_with_ledger_title(self):
        ledger = pd.DataFrame(
            [
                {
                    "platform": "B站",
                    "content_id": "BV1abcde2345",
                    "content_url": "https://www.bilibili.com/video/BV1abcde2345/",
                    "title": "真实B站标题",
                    "account": "投资号",
                    "content_type": "采访内容",
                    "bilibili_content_type": "采访内容",
                }
            ]
        )
        frame = normalize_platform_identities(
            pd.DataFrame(
                [
                    {
                        "platform": "B站",
                        "channel": "B站市场部",
                        "material_id": "BV1abcde2345",
                        "title": "BV1abcde2345",
                        "spend": 100,
                        "impressions": 1000,
                    }
                ]
            )
        )

        matched = match_assets_to_ledger(frame, ledger)

        self.assertEqual(matched.iloc[0]["match_status"], "已匹配")
        self.assertEqual(matched.iloc[0]["match_source"], "BV号")
        self.assertEqual(matched.iloc[0]["title"], "真实B站标题")
        self.assertEqual(matched.iloc[0]["matched_ledger_title"], "真实B站标题")

    def test_analysis_scope_and_core_tables_follow_three_status_contract(self):
        matched = match_assets_to_ledger(
            normalize_platform_identities(
                pd.DataFrame(
                    [
                        {
                            "period_start": "2026-05-01",
                            "period_end": "2026-05-07",
                            "platform": "小红书",
                            "channel": "小红书商业化",
                            "title": "投放标题",
                            "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                            "spend": 100,
                            "impressions": 1000,
                            "source_file": "xhs.xlsx",
                            "source_row": 2,
                        },
                        {
                            "period_start": "2026-05-01",
                            "period_end": "2026-05-07",
                            "platform": "抖音",
                            "channel": "抖音商业化",
                            "title": "A组-推送素材",
                            "content_url": "https://v.douyin.com/unresolved/",
                            "material_id": "7391234567890123456",
                            "spend": 200,
                            "impressions": 2000,
                            "source_file": "douyin.xlsx",
                            "source_row": 3,
                        },
                        {
                            "period_start": "2026-05-01",
                            "period_end": "2026-05-07",
                            "platform": "微信",
                            "channel": "微信",
                            "title": "其他平台",
                            "spend": 300,
                            "impressions": 3000,
                            "source_file": "wechat.xlsx",
                            "source_row": 4,
                        },
                    ]
                )
            ),
            _ledger(),
        )
        scoped = apply_analysis_scope(matched)

        self.assertEqual(list(scoped["analysis_status"]), ["可分析", "待补全", "不可分析"])
        self.assertEqual(scoped.iloc[1]["unanalyzable_reason"], "抖音URL解析失败")
        self.assertEqual(scoped.iloc[2]["unanalyzable_reason"], "平台不在复盘范围")

        asset_table = build_cleaned_asset_table(scoped)
        recap_table = build_content_recap_table(asset_table)
        summary = build_unanalyzable_summary(asset_table)

        self.assertEqual(list(asset_table.columns), CLEANED_ASSET_COLUMNS)
        self.assertEqual(list(recap_table.columns), CONTENT_RECAP_COLUMNS)
        self.assertEqual(list(summary.columns), UNANALYZABLE_SUMMARY_COLUMNS)
        self.assertEqual(len(recap_table), 1)
        self.assertEqual(recap_table.iloc[0]["内容类型"], "资讯")
        self.assertEqual(summary["不可分析素材数"].sum(), 2)
        self.assertIn("抖音URL解析失败", set(summary["主要原因"]))

    def test_core_export_contains_business_and_attribution_sheets(self):
        scoped = apply_analysis_scope(
            match_assets_to_ledger(
                normalize_platform_identities(
                    pd.DataFrame(
                        [
                            {
                                "period_start": "2026-05-01",
                                "period_end": "2026-05-07",
                                "platform": "B站",
                                "channel": "B站市场部",
                                "material_id": "BV1abcde2345",
                                "title": "B站投放",
                                "spend": 100,
                                "impressions": 1000,
                            }
                        ]
                    )
                ),
                _ledger(),
            )
        )
        asset_table = build_cleaned_asset_table(scoped)
        recap_table = build_content_recap_table(asset_table)
        summary = build_unanalyzable_summary(asset_table)

        with TemporaryDirectory() as tmp:
            path = write_core_recap_workbook(Path(tmp), asset_table, recap_table, summary)
            workbook = load_workbook(path, read_only=True)

        self.assertEqual(path.name, CORE_RECAP_WORKBOOK)
        self.assertEqual(workbook.sheetnames, CORE_ANALYSIS_SHEETS)

    def test_archived_workflow_writes_only_core_recap_outputs_and_core_sqlite_tables(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            raw_dir = root / "raw"
            raw_dir.mkdir()
            with pd.ExcelWriter(raw_dir / "小红书商业化.xlsx", engine="openpyxl") as writer:
                pd.DataFrame(
                    [
                        {
                            "标题": "投放标题",
                            "笔记链接": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                            "消费": 100,
                            "展现量": 1000,
                        }
                    ]
                ).to_excel(writer, index=False)
            with patch("ops_data_workflow.raw_cleaning.load_feishu_content_ledger", return_value=_ledger()):
                result = run_archived_workflow(
                    raw_dir,
                    "2026-05-01",
                    "2026-05-07",
                    output_root=root / "outputs",
                    processed_root=root / "processed",
                    db_path=root / "workflow.sqlite3",
                    enable_deepseek=False,
                    enable_external_context=False,
                    metadata_enrichment_mode="off",
                )

            self.assertTrue(result.core_recap_xlsx.exists())
            workbook = load_workbook(result.core_recap_xlsx, read_only=True)
            self.assertEqual(workbook.sheetnames, CORE_ANALYSIS_SHEETS)

            with sqlite3.connect(root / "workflow.sqlite3") as conn:
                asset_count = conn.execute(
                    "select count(*) from cleaned_asset_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                recap_count = conn.execute(
                    "select count(*) from content_recap_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
                summary_count = conn.execute(
                    "select count(*) from unanalyzable_summary_items where batch_id = ?",
                    (result.batch_id,),
                ).fetchone()[0]
            self.assertEqual(asset_count, 1)
            self.assertEqual(recap_count, 1)
            self.assertEqual(summary_count, 1)

    def test_archived_workflow_matches_period_against_refreshed_local_asset_table(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            db_path = root / "workflow.sqlite3"
            first_raw = root / "first"
            second_raw = root / "second"
            first_raw.mkdir()
            second_raw.mkdir()
            pd.DataFrame(
                [
                    {
                        "标题": "第一次上传占位标题",
                        "笔记链接": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                        "消费": 10,
                        "展现量": 100,
                    }
                ]
            ).to_excel(first_raw / "小红书商业化.xlsx", index=False)
            pd.DataFrame(
                [
                    {
                        "标题": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                        "笔记链接": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                        "消费": 20,
                        "展现量": 200,
                    }
                ]
            ).to_excel(second_raw / "小红书商业化.xlsx", index=False)
            full_ledger = _ledger()
            blank_refresh = pd.DataFrame(
                [
                    {
                        "platform": "小红书",
                        "content_id": "65f00000abcdef",
                        "content_url": "https://www.xiaohongshu.com/explore/65f00000abcdef",
                        "title": "",
                        "account": "",
                        "content_type": "",
                        "source_file": "harvester_feishu",
                        "source_sheet": "小红书",
                        "source_row": 2,
                    }
                ]
            )

            with patch(
                "ops_data_workflow.raw_cleaning.load_feishu_content_ledger",
                side_effect=[full_ledger, full_ledger, blank_refresh, blank_refresh],
            ):
                run_archived_workflow(
                    first_raw,
                    "2026-05-01",
                    "2026-05-07",
                    output_root=root / "outputs",
                    processed_root=root / "processed",
                    db_path=db_path,
                    enable_deepseek=False,
                    enable_external_context=False,
                    metadata_enrichment_mode="off",
                )
                result = run_archived_workflow(
                    second_raw,
                    "2026-05-08",
                    "2026-05-14",
                    output_root=root / "outputs",
                    processed_root=root / "processed",
                    db_path=db_path,
                    enable_deepseek=False,
                    enable_external_context=False,
                    metadata_enrichment_mode="off",
                )

            row = result.canonical.iloc[0]
            self.assertEqual(row["match_status"], "已匹配")
            self.assertEqual(row["matched_ledger_title"], "小红书自有标题")
            self.assertEqual(row["title"], "小红书自有标题")
            self.assertEqual(row["account"], "投资号")
            self.assertEqual(row["category_l2"], "资讯")
