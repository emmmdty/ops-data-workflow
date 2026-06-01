from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook
import pandas as pd
import yaml

from ops_data_workflow.field_mapping import (
    load_field_mapping,
    mapped_or_ignored_headers,
    standardize_content_form,
    standardize_content_type,
)


def _default_config() -> dict:
    config_path = Path(__file__).resolve().parents[1] / "config" / "field_mapping.yml"
    return yaml.safe_load(config_path.read_text(encoding="utf-8"))


def _write_config(data: dict, root: Path) -> Path:
    path = root / "field_mapping.yml"
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return path


def _local_data_headers() -> set[str]:
    base = Path(__file__).resolve().parents[2] / "data"
    if not base.exists():
        raise unittest.SkipTest("../data is not available in this checkout")
    headers: set[str] = set()
    for path in sorted(base.rglob("*")):
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        if path.name.startswith(".~") or path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        headers.update(values)
                        break
        finally:
            workbook.close()
    return headers


class FieldMappingTests(unittest.TestCase):
    def _assert_invalid_config(self, data: dict, message: str) -> None:
        with TemporaryDirectory() as tmp:
            path = _write_config(data, Path(tmp))
            with self.assertRaisesRegex(ValueError, message):
                load_field_mapping(path)

    def test_config_loads_with_unique_internal_fields(self):
        mapping = load_field_mapping()

        internal_fields = [field.internal for field in mapping.fields]
        self.assertEqual(len(internal_fields), len(set(internal_fields)))
        self.assertIn("content_form", internal_fields)
        self.assertIn("manual_category", internal_fields)
        self.assertNotIn("类型", mapping.field_by_internal("manual_category").source_columns)

    def test_config_validation_rejects_missing_required_field(self):
        data = _default_config()
        data["fields"] = [field for field in data["fields"] if field["internal"] != "title"]

        self._assert_invalid_config(data, "missing required fields")

    def test_config_validation_rejects_invalid_role(self):
        data = _default_config()
        data["fields"][0]["role"] = "legacy_alias"

        self._assert_invalid_config(data, "invalid role")

    def test_config_validation_rejects_duplicate_source_in_field(self):
        data = _default_config()
        first_source = data["fields"][0]["source_columns"][0]
        data["fields"][0]["source_columns"].append(first_source)

        self._assert_invalid_config(data, "duplicate source columns")

    def test_config_validation_rejects_ignored_mapped_conflict(self):
        data = _default_config()
        data["ignored_fields"].append(data["fields"][0]["source_columns"][0])

        self._assert_invalid_config(data, "both mapped and ignored")

    def test_internal_passthrough_fields_are_configured_but_not_exported(self):
        mapping = load_field_mapping()
        fields = mapping.fields_for_source("generic")

        self.assertEqual(fields["category_l3"], ["三级题材", "题材"])
        self.assertEqual(fields["category_status"], ["类别来源_解析"])
        self.assertNotIn("category_l3", set(mapping.to_frame()["标准字段"]))
        self.assertNotIn("category_status", set(mapping.to_frame()["标准字段"]))

    def test_local_data_headers_are_mapped_or_ignored(self):
        mapping = load_field_mapping()
        headers = _local_data_headers()

        result = mapped_or_ignored_headers(headers, mapping)

        self.assertEqual(result.unmapped, set())
        self.assertGreater(len(result.mapped), 0)
        self.assertGreater(len(result.ignored), 0)

    def test_type_only_sets_content_form_without_backfilling_content_type(self):
        row = pd.Series({"类型": "图文", "内容分类": "", "内容类型": ""})

        self.assertEqual(standardize_content_form(row), "图文")
        self.assertEqual(standardize_content_type(row), "")

    def test_graphic_content_type_sets_form_and_type(self):
        for column in ["内容类型", "内容分类"]:
            row = pd.Series({"类型": "", "内容分类": "", "内容类型": "", column: "图文"})

            self.assertEqual(standardize_content_form(row), "图文")
            self.assertEqual(standardize_content_type(row), "图文")

    def test_non_graphic_content_type_defaults_form_to_video(self):
        row = pd.Series({"类型": "", "内容分类": "", "内容类型": "股友说"})

        self.assertEqual(standardize_content_form(row), "视频")
        self.assertEqual(standardize_content_type(row), "股友说")

    def test_content_category_takes_priority_over_content_type(self):
        row = pd.Series({"类型": "图文", "内容分类": "互动话题", "内容类型": "图文"})

        self.assertEqual(standardize_content_form(row), "图文")
        self.assertEqual(standardize_content_type(row), "互动话题")

    def test_invalid_content_type_values_are_ignored(self):
        for value in ["", "#REF!", "0"]:
            row = pd.Series({"类型": "视频", "内容分类": "", "内容类型": value})
            self.assertEqual(standardize_content_type(row), "")


if __name__ == "__main__":
    unittest.main()
